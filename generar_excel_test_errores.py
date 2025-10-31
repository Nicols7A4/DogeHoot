import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Crear workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Preguntas con Errores"

# Estilo de encabezados
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
header_alignment = Alignment(horizontal="center", vertical="center")

# Encabezados
headers = ['pregunta', 'opcion_1', 'opcion_2', 'opcion_3', 'opcion_4', 'opcion_c_orrecta', 'tiempo_segundos', 'url_imagen']
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Datos de prueba CON ERRORES INTENCIONALES
preguntas = [
    # Fila 2: ✅ VÁLIDA (para comparar)
    {
        'pregunta': '¿Qué es un Flask?',
        'opcion_1': 'Un Framework de desarrollo web',
        'opcion_2': 'Un lenguaje de programación',
        'opcion_3': 'Un Framework de desarrollo de escritorio',
        'opcion_4': 'Una base de datos',
        'opcion_c_orrecta': 1,
        'tiempo_segundos': 15,
        'url_imagen': ''
    },
    # Fila 3: ❌ ERROR - Tiempo menor a 10
    {
        'pregunta': '¿Qué día es hoy?',
        'opcion_1': 'Martes',
        'opcion_2': 'Viernes',
        'opcion_3': 'Jueves',
        'opcion_4': 'Lunes',
        'opcion_c_orrecta': 3,
        'tiempo_segundos': 5,  # ❌ ERROR: menor que 10
        'url_imagen': ''
    },
    # Fila 4: ❌ ERROR - Opción correcta vacía
    {
        'pregunta': 'Fecha de nacimiento',
        'opcion_1': 'Ayer',
        'opcion_2': 'Hoy',
        'opcion_3': 'Nosé',
        'opcion_4': '',  # ❌ Opción 4 está vacía
        'opcion_c_orrecta': 4,  # ❌ ERROR: apunta a una opción vacía
        'tiempo_segundos': 10,
        'url_imagen': ''
    },
    # Fila 5: ❌ ERROR - Solo 1 opción con texto
    {
        'pregunta': '¿Cuánto es 2 + 2?',
        'opcion_1': '4',
        'opcion_2': '',
        'opcion_3': '',
        'opcion_4': '',
        'opcion_c_orrecta': 1,
        'tiempo_segundos': 20,  # ❌ ERROR: solo tiene 1 opción
        'url_imagen': ''
    },
    # Fila 6: VACÍA (debería ignorarse)
    {},
    # Fila 7: ❌ ERROR - Pregunta vacía
    {
        'pregunta': '',  # ❌ ERROR: pregunta vacía
        'opcion_1': 'Opción A',
        'opcion_2': 'Opción B',
        'opcion_3': 'Opción C',
        'opcion_4': 'Opción D',
        'opcion_c_orrecta': 2,
        'tiempo_segundos': 15,
        'url_imagen': ''
    },
    # Fila 8: ❌ ERROR - opcion_c_orrecta fuera de rango
    {
        'pregunta': '¿Cuál es la capital de Francia?',
        'opcion_1': 'Londres',
        'opcion_2': 'París',
        'opcion_3': 'Berlín',
        'opcion_4': 'Madrid',
        'opcion_c_orrecta': 5,  # ❌ ERROR: debe ser 1-4
        'tiempo_segundos': 15,
        'url_imagen': ''
    },
    # Fila 9: ❌ ERROR - tiempo_segundos con texto
    {
        'pregunta': '¿Qué es Python?',
        'opcion_1': 'Un lenguaje',
        'opcion_2': 'Una serpiente',
        'opcion_3': 'Un framework',
        'opcion_4': 'Una base de datos',
        'opcion_c_orrecta': 1,
        'tiempo_segundos': 'quince',  # ❌ ERROR: texto en lugar de número
        'url_imagen': ''
    },
    # Fila 10: ❌ ERROR - tiempo > 100
    {
        'pregunta': '¿Cuántos días tiene un año?',
        'opcion_1': '365',
        'opcion_2': '366',
        'opcion_3': '364',
        'opcion_4': '360',
        'opcion_c_orrecta': 1,
        'tiempo_segundos': 150,  # ❌ ERROR: mayor que 100
        'url_imagen': ''
    },
    # Fila 11: ✅ VÁLIDA - con URL de imagen
    {
        'pregunta': '¿Qué es GitHub?',
        'opcion_1': 'Un repositorio',
        'opcion_2': 'Un lenguaje',
        'opcion_3': 'Una base de datos',
        'opcion_4': '',
        'opcion_c_orrecta': 1,
        'tiempo_segundos': 20,
        'url_imagen': 'https://example.com/github.png'
    },
]

# Insertar datos
for row_idx, pregunta in enumerate(preguntas, start=2):
    if not pregunta:  # Fila vacía
        continue
    
    ws.cell(row=row_idx, column=1, value=pregunta.get('pregunta', ''))
    ws.cell(row=row_idx, column=2, value=pregunta.get('opcion_1', ''))
    ws.cell(row=row_idx, column=3, value=pregunta.get('opcion_2', ''))
    ws.cell(row=row_idx, column=4, value=pregunta.get('opcion_3', ''))
    ws.cell(row=row_idx, column=5, value=pregunta.get('opcion_4', ''))
    ws.cell(row=row_idx, column=6, value=pregunta.get('opcion_c_orrecta', ''))
    ws.cell(row=row_idx, column=7, value=pregunta.get('tiempo_segundos', ''))
    ws.cell(row=row_idx, column=8, value=pregunta.get('url_imagen', ''))

# Ajustar anchos de columna
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 35
ws.column_dimensions['E'].width = 35
ws.column_dimensions['F'].width = 20
ws.column_dimensions['G'].width = 18
ws.column_dimensions['H'].width = 40

# Guardar archivo
wb.save('test_importacion_con_errores.xlsx')
print("✅ Archivo de prueba 'test_importacion_con_errores.xlsx' creado exitosamente!")
print("\nErrores intencionales incluidos:")
print("- Fila 3: Tiempo = 5 (menor que 10)")
print("- Fila 4: Opción correcta (4) está vacía")
print("- Fila 5: Solo 1 opción con texto")
print("- Fila 6: Fila completamente vacía (debe ignorarse)")
print("- Fila 7: Pregunta vacía")
print("- Fila 8: opcion_c_orrecta = 5 (fuera de rango)")
print("- Fila 9: tiempo_segundos = 'quince' (texto)")
print("- Fila 10: tiempo_segundos = 150 (mayor que 100)")
print("\nFilas válidas: 2 y 11")
