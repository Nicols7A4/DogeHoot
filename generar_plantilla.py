import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Crear libro de trabajo
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Preguntas'

# Definir encabezados con el nuevo formato (incluye tiempo_segundos)
headers = ['pregunta', 'opcion_1', 'opcion_2', 'opcion_3', 'opcion_4', 'opcion_c_orrecta', 'tiempo_segundos', 'url_imagen']

# Estilos para encabezados
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
header_alignment = Alignment(horizontal='center', vertical='center')

# Escribir encabezados
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Ejemplos de preguntas (ahora con tiempo_segundos)
ejemplo1 = ['Que dia es hoy?', 'Martes', 'Viernes', 'Jueves', 'Lunes', 3, 15, '']
ejemplo2 = ['2 + 2', '4', '5', '0', '1', 1, 20, '']

# Escribir ejemplos
for col_num, value in enumerate(ejemplo1, 1):
    ws.cell(row=2, column=col_num, value=value)

for col_num, value in enumerate(ejemplo2, 1):
    ws.cell(row=3, column=col_num, value=value)

# Ajustar ancho de columnas
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column].width = adjusted_width

# Guardar
wb.save('mysite/static/cuestionario_plantilla.xlsx')
print('✅ Plantilla Excel creada exitosamente en: mysite/static/cuestionario_plantilla.xlsx')
print('📋 Formato: pregunta | opcion_1 | opcion_2 | opcion_3 | opcion_4 | opcion_c_orrecta (1-4) | tiempo_segundos (10-100) | url_imagen (opcional)')
