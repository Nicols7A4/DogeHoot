from flask import jsonify, request, url_for, session, Response, send_file, redirect, current_app
import traceback
import os
import time
from main import app
from werkzeug.utils import secure_filename
import csv          #usado para los reportes
import io           #usado para los reportes
import pymysql
from datetime import datetime
import hashlib

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Border, Side
from openpyxl.utils import get_column_letter

from bd import obtener_conexion
from controladores import cuestionarios as cc
from controladores import preguntas_opciones as cpo
from controladores import controlador_partidas as c_part
from controladores import controlador_recompensas as c_rec
from controladores import usuarios as ctrl
from controladores import outlook_email_sender as email_sender
from controladores.onedrive_uploader import OneDriveUploader
from ajax_game import (
    partidas_en_juego,
    _ensure_loaded,
    get_lobby_state,
    join_player,
    select_group,
    remove_player,
    start_game,
    get_status,
    get_current,
    submit_answer,
    advance_state_if_needed,
    finalize_game,
)

from flask_jwt import JWT, jwt_required

# --- Auxiliares --------
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ----------------------------


def encriptar_sha256(texto):
    texto = texto.encode('utf-8')
    objHash = hashlib.sha256(texto)
    textenc = objHash.hexdigest()
    return textenc


@app.route("/api/test")
def api_test():
    
    return jsonify({"mensaje": "La API funciona correctamente"})


# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# USUARIOS
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------


@app.route("/api_obtener_usuarios", methods=['GET'])
@jwt_required()
def fn_api_obtener_usuarios():
    try:
        usuarios = ctrl.obtener_todos()
        return jsonify({"data":usuarios}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500 

@app.route("/api_obtener_usuario_por_id/<int:id_usuario>", methods=['GET'])
@jwt_required()
def fn_api_obtener_usuario_por_id(id_usuario):
    try:
        # id_usuario = request.args.get('id_usuario', type=int)
        usuario = ctrl.obtener_por_id(id_usuario)
        return jsonify({"data":usuario}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500 

@app.route("/api_registrar_usuario", methods=['POST'])
@jwt_required()
def fn_api_registrar_usuario():
    data = request.get_json(force=True) or {}
    nombre_completo = (data.get("nombre_completo") or "").strip()
    nombre_usuario  = (data.get("nombre_usuario")  or "").strip()
    correo          = (data.get("correo")          or "").strip().lower()
    contrasena      = (data.get("contrasena")      or "")
    tipo            = (data.get("tipo")            or "E").strip()

    conexion = obtener_conexion()
    try:
        with conexion.cursor() as c:
            c.execute("SELECT id_usuario FROM USUARIO WHERE correo=%s AND vigente = true;", (correo))
            if c.fetchone():
                return jsonify({"error":"El correo ya está en uso", "campo":"correo"}), 409
            c.execute("SELECT id_usuario FROM USUARIO WHERE nombre_usuario=%s AND vigente = true;", (nombre_usuario))
            if c.fetchone():
                return jsonify({"error":"El nombre de usuario ya está en uso", "campo":"nombre_usuario"}), 409
    finally:
        if conexion:
            conexion.close()

    if not (nombre_completo and nombre_usuario and correo and contrasena):
        return jsonify({"error": "Faltan campos obligatorios"}), 400
    

    if not (correo.endswith("@usat.edu.pe") or correo.endswith("@usat.pe")):
        return jsonify({"error": "Solo son válidas correos de la USAT"}), 400
    
    if len(contrasena) < 8:
        return jsonify({"error": "La contraseña debe tener como mínimo 8 caractares"}), 400
    if not any(c.islower() for c in contrasena):
        return jsonify({"error": "La contraseña debe contener al menos una letra minúscula"}), 400
    if not any(c.isupper() for c in contrasena):
        return jsonify({"error": "La contraseña debe contener al menos una letra mayúscula"}), 400
    if not any(c.isdigit() for c in contrasena):
        return jsonify({"error": "La contraseña debe contener al menos un número"}), 400
    if not any(c in "!@#$%^&*()_+-=[]{}|;:,.<>?/~`" for c in contrasena):
        return jsonify({"error": "La contraseña debe contener al menos un símbolo especial"}), 400
    
    if tipo != "P" or tipo != "E":
            return jsonify({"error": "Tipo de usuario inválido"}), 404
    tipo = "P" if correo.endswith("@usat.edu.pe") else "E"

    modo = (request.args.get("modo") or "pendiente").lower()
    if modo == "directo" or True:
        ok, msg = ctrl.crear_usuario_t(nombre_completo, nombre_usuario, correo, contrasena, tipo)
        return (jsonify({"ok": True, "mensaje": msg}), 201) if ok else (jsonify({"error": msg}), 409)

    # ok, resultado = ctrl.crear_usuario_pendiente(nombre_completo, nombre_usuario, correo, contrasena, tipo)
    # if ok:
    #     return jsonify({"ok": True, "Mensaje":"Usuario registrado"}), 201
    # return jsonify({"error": resultado}), 409

@app.route("/api_actualizar_usuario", methods=['PUT'])
@jwt_required()
def fn_api_actulizar_usuario():
    try:
        data = request.get_json(force=True) or {}
        id_usuario = data.get("id_usuario")
        # Lógica para actualizar un usuario
        actual = ctrl.obtener_por_id(id_usuario)
        if not actual:
            return jsonify({"error": "Usuario no encontrado"}), 404

        data = request.get_json(force=True) or {}

        # merge con valores actuales
        nombre_completo = (data.get("nombre_completo") or actual["nombre_completo"]).strip()
        nombre_usuario  = (data.get("nombre_usuario")  or actual["nombre_usuario"]).strip()
        correo          = (data.get("correo")          or actual["correo"]).strip().lower()
        tipo            = (data.get("tipo")            or actual["tipo"]).strip()
        puntos          = int(data.get("puntos")  if data.get("puntos")  is not None else actual["puntos"])
        monedas         = int(data.get("monedas") if data.get("monedas") is not None else actual["monedas"])
        vigente         = bool(data.get("vigente") if data.get("vigente") is not None else actual["vigente"])

        if tipo != "P" or tipo != "E":
            return jsonify({"error": "Tipo de usuario inválido"}), 404
        if monedas <= 0:
            return jsonify({"error": "Cantidad de monedas inválida"}), 404
        if puntos <= 0:
            return jsonify({"error": "Cantidad de puntos inválida"}), 404
        
        
        # unicidad correo / nombre_usuario (excluyendo mi propio id)
        conexion = obtener_conexion()
        try:
            with conexion.cursor() as c:
                c.execute("SELECT id_usuario FROM USUARIO WHERE correo=%s AND id_usuario<>%s LIMIT 1", (correo, id_usuario))
                if c.fetchone():
                    return jsonify({"error":"El correo ya está en uso", "campo":"correo"}), 409
                c.execute("SELECT id_usuario FROM USUARIO WHERE nombre_usuario=%s AND id_usuario<>%s LIMIT 1", (nombre_usuario, id_usuario))
                if c.fetchone():
                    return jsonify({"error":"El nombre de usuario ya está en uso", "campo":"nombre_usuario"}), 409
        finally:
            if conexion:
                conexion.close()

        ctrl.actualizar(id_usuario, nombre_completo, nombre_usuario, correo, tipo, puntos, monedas, vigente)
        actualizado = ctrl.obtener_por_id(id_usuario)
        actualizado.pop("contraseña", None)
        return jsonify(actualizado), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500 

@app.route("/api_eliminar_usuario", methods=['POST','DELETE'])
@jwt_required()
def fn_api_eliminar_usuario():
    try:
        data = request.get_json(force=True) or {}
        id_usuario = data.get("id_usuario")
        # Lógica para eliminar (o desactivar) un usuario
        if not ctrl.obtener_por_id(id_usuario):
            return jsonify({"error": "Usuario no encontrado"}), 404

        ok = ctrl.desactivar(id_usuario)  # True/False
        if not ok:
            # No se afectó ninguna fila
            return jsonify({"error": "No se pudo desactivar el usuario"}), 500

        return jsonify({"ok": True, "id_usuario": id_usuario, "vigente": False}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500 


# ---------------------------------------------------------------

@app.route("/api/usuarios", methods=['POST'])
def api_create_usuario():
    data = request.get_json(force=True) or {}
    nombre_completo = (data.get("nombre_completo") or "").strip()
    nombre_usuario  = (data.get("nombre_usuario")  or "").strip()
    correo          = (data.get("correo")          or "").strip().lower()
    contrasena      = (data.get("contrasena")      or "")
    tipo            = (data.get("tipo")            or "E").strip()

    if not (nombre_completo and nombre_usuario and correo and contrasena):
        return jsonify({"error": "Faltan campos obligatorios"}), 400

    modo = (request.args.get("modo") or "pendiente").lower()
    if modo == "directo":
        ok, msg = ctrl.crear_usuario(nombre_completo, nombre_usuario, correo, contrasena, tipo)
        return (jsonify({"ok": True, "mensaje": msg}), 201) if ok else (jsonify({"error": msg}), 409)

    ok, resultado = ctrl.crear_usuario_pendiente(nombre_completo, nombre_usuario, correo, contrasena, tipo)
    if ok:
        return jsonify({"ok": True, "codigo_verificacion": resultado}), 201
    return jsonify({"error": resultado}), 409

@app.route("/api/usuarios/verificacion", methods=["POST"])
def usuarios_verificar():
    data = request.get_json(force=True) or {}
    correo = (data.get("correo") or "").strip().lower()
    codigo = (data.get("codigo") or "").strip()
    if not (correo and codigo):
        return jsonify({"error": "correo y codigo son obligatorios"}), 400

    res = ctrl.verificar_y_activar_usuario(correo, codigo)
    if res == "OK":
        return jsonify({"ok": True}), 200
    return jsonify({"error": res}), 400

@app.route("/api/usuarios/<int:id_usuario>", methods=['GET'])
def api_get_usuario(id_usuario):
    u = ctrl.obtener_por_id(id_usuario)
    if not u:
        return jsonify({"error": "Usuario no encontrado"}), 404
    # no exponer contraseña
    u.pop("contraseña", None)
    return jsonify(u), 200


@app.route("/api/usuarios/<int:id_usuario>", methods=['PUT'])
def api_update_usuario(id_usuario):
    # Lógica para actualizar un usuario
    actual = ctrl.obtener_por_id(id_usuario)
    if not actual:
        return jsonify({"error": "Usuario no encontrado"}), 404

    data = request.get_json(force=True) or {}

    # merge con valores actuales
    nombre_completo = (data.get("nombre_completo") or actual["nombre_completo"]).strip()
    nombre_usuario  = (data.get("nombre_usuario")  or actual["nombre_usuario"]).strip()
    correo          = (data.get("correo")          or actual["correo"]).strip().lower()
    tipo            = (data.get("tipo")            or actual["tipo"]).strip()
    puntos          = int(data.get("puntos")  if data.get("puntos")  is not None else actual["puntos"])
    monedas         = int(data.get("monedas") if data.get("monedas") is not None else actual["monedas"])
    vigente         = bool(data.get("vigente") if data.get("vigente") is not None else actual["vigente"])

    # unicidad correo / nombre_usuario (excluyendo mi propio id)
    conexion = obtener_conexion()
    try:
        with conexion.cursor() as c:
            c.execute("SELECT id_usuario FROM USUARIO WHERE correo=%s AND id_usuario<>%s LIMIT 1", (correo, id_usuario))
            if c.fetchone():
                return jsonify({"error":"El correo ya está en uso", "campo":"correo"}), 409
            c.execute("SELECT id_usuario FROM USUARIO WHERE nombre_usuario=%s AND id_usuario<>%s LIMIT 1", (nombre_usuario, id_usuario))
            if c.fetchone():
                return jsonify({"error":"El nombre de usuario ya está en uso", "campo":"nombre_usuario"}), 409
    finally:
        if conexion:
            conexion.close()

    ctrl.actualizar(id_usuario, nombre_completo, nombre_usuario, correo, tipo, puntos, monedas, vigente)
    actualizado = ctrl.obtener_por_id(id_usuario)
    actualizado.pop("contraseña", None)
    return jsonify(actualizado), 200

@app.route("/api/usuarios/<int:id_usuario>/cambiar_password", methods=['POST'])
def api_cambiar_password(id_usuario):
    try:
        data = request.get_json(force=True) or {}
    except Exception:
        return jsonify({"error": "Solicitud inválida"}), 400

    antigua = (data.get("actual") or "").strip()
    nueva   = (data.get("nueva")  or "").strip()

    if not antigua or not nueva:
        return jsonify({"error": "Faltan datos"}), 400

    ok, msg = ctrl.actualizar_contrasena(id_usuario, antigua, nueva)

    if ok:
        return jsonify({"ok": True}), 200

    # Mapeo de mensajes a lo que espera el frontend
    m = (msg or "").lower()
    if "usuario no encontrado" in m:
        return jsonify({"error": "Usuario no encontrado"}), 404
    if "contraseña actual es incorrecta" in m or "contrasena actual es incorrecta" in m:
        return jsonify({"error": "Tu contraseña actual no es la correcta"}), 409

    return jsonify({"error": "Algo salió mal, vuelve a intentarlo"}), 500

@app.route("/api/usuarios/<int:id_usuario>", methods=['DELETE'])
def api_delete_usuario(id_usuario):
    # Lógica para eliminar (o desactivar) un usuario
    if not ctrl.obtener_por_id(id_usuario):
        return jsonify({"error": "Usuario no encontrado"}), 404

    ok = ctrl.desactivar(id_usuario)  # True/False
    if not ok:
        # No se afectó ninguna fila
        return jsonify({"error": "No se pudo desactivar el usuario"}), 500

    return jsonify({"ok": True, "id_usuario": id_usuario, "vigente": False}), 200

@app.route("/api/verificar_correo/<correo>")
def verificar_correo(correo):
    try:
        conn = obtener_conexion()
        cursor = conn.cursor(dictionary=True)
        query = "SELECT vigente FROM USUARIO WHERE correo = %s"
        cursor.execute(query, (correo,))
        usuario = cursor.fetchone()
        cursor.close()
        conn.close()

        if usuario:
            return jsonify({"existe": True, "vigente": bool(usuario["vigente"])})
        else:
            return jsonify({"existe": False, "vigente": False})
    except Exception as e:
        print("Error en verificar_correo:", e)
        return jsonify({"error": "Error interno"}), 500


@app.route("/api/correo_activo", methods=["GET"])
def api_correo_activo():
    correo = (request.args.get("correo") or "").strip()
    if not correo:
        return jsonify({"ok": False, "error": "correo_requerido"}), 400

    require_verificado = request.args.get("require_verificado", "false").lower() == "true"

    try:
        valido = ctrl.correo_activo(correo, require_verificado=require_verificado)
        return jsonify({"ok": True, "valido": bool(valido), "require_verificado": require_verificado}), 200
    except Exception:
        traceback.print_exc()  # <-- log más útil
        return jsonify({"ok": False, "error": "interno"}), 500


# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# CUESTIONARIOS Y SUS PARTES
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------

@app.route("/api_obtener_cuestionarios", methods=['GET'])
@jwt_required()
def fn_api_obtener_cuestionarios():
    try:
        cuestionarios = cc.obtener_cuestionarios_todos()
        return jsonify({"data":cuestionarios}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500 

@app.route("/api_obtener_cuestionario_por_id/<int:id_cuestionario>", methods=['GET'])
@jwt_required()
def fn_api_obtener_cuestionario_por_id(id_cuestionario):
    try:
        cuestionario = cc.obtener_cuestionario_por_id(id_cuestionario)
        return jsonify({"data":cuestionario}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500 

@app.route("/api_registar_cuestionario", methods=['POST'])
@jwt_required()
def fn_api_registrar_cuestionario():
    try:
        data = request.json
        cc.crear(
            titulo=data['titulo'],
            descripcion=data['descripcion'],
            es_publico=data['es_publico'],
            id_usuario=data['id_usuario'],
            id_categoria=data['id_categoria'],
            id_cuestionario_original=data.get('id_cuestionario_original') # .get() para campos opcionales
        )
        return jsonify({"mensaje": "Cuestionario creado con éxito"}), 201
    except KeyError:
        return jsonify({"error": "Faltan datos requeridos"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api_actualizar_cuestionario", methods=['PUT'])
@jwt_required()
def fn_api_actualizar_cuestionario():
    try:
        data = request.json
        cc.actualizar(
            id_cuestionario=data['id_cuestionario'],
            titulo=data['titulo'],
            descripcion=data['descripcion'],
            es_publico=data['es_publico'],
            #id_usuario=data['id_usuario'],
            id_categoria=data['id_categoria'],
            vigente=data['vigente']
        )
        return jsonify({"mensaje": "Cuestionario actualizado con éxito"}), 201
    except KeyError:
        return jsonify({"error": "Faltan datos requeridos"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@jwt_required()
def fn_api_eliminar_cuestionario():
    try:
        data = request.get_json(force=True) or {}
        id_cuestionario = data.get("id_cuestionario")
        # Llama a la función del controlador para desactivar
        exito = cc.desactivar(id_cuestionario)

        if exito:
            return jsonify({"mensaje": "Cuestionario desactivado correctamente"}), 200
        else:
            return jsonify({"error": "Cuestionario no encontrado"}), 404

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------

@app.route("/api/obtener_cuestionarios", methods=['GET'])
def api_obtenercuestionarios():
    try:
        cuestionarios = cc.obtener_cuestionarios_todos()
        return jsonify({"data":cuestionarios}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500 

@app.route("/api/cuestionarios", methods=['GET'])
def api_get_cuestionarios():
    try:
        # Leemos los filtros opcionales de la URL
        id_usuario = request.args.get('id_usuario', type=int)
        id_categoria = request.args.get('id_categoria', type=int)
        es_publico = request.args.get('publico', type=bool)

        # Llamamos al controlador con los filtros
        cuestionarios = cc.obtener_con_filtros(id_usuario, id_categoria, es_publico)
        return jsonify({"data": cuestionarios}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/cuestionarios", methods=['POST'])
def api_create_cuestionario():
    try:
        data = request.json
        cc.crear(
            titulo=data['titulo'],
            descripcion=data['descripcion'],
            es_publico=data['es_publico'],
            id_usuario=data['id_usuario'],
            id_categoria=data['id_categoria'],
            id_cuestionario_original=data.get('id_cuestionario_original') # .get() para campos opcionales
        )
        return jsonify({"mensaje": "Cuestionario creado con éxito"}), 201
    except KeyError:
        return jsonify({"error": "Faltan datos requeridos"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/cuestionarios/<int:id_cuestionario>", methods=['GET'])
def api_get_cuestionario(id_cuestionario):
    cuestionario = cc.obtener_por_id(id_cuestionario)
    if cuestionario:
        return jsonify({"data": cuestionario}), 200
    return jsonify({"mensaje": "Cuestionario no encontrado"}), 404

@app.route("/api/cuestionarios", methods=['PUT'])
def api_update_cuestionario():
    try:
        data = request.json
        cc.actualizar(
            id_cuestionario=data['id_cuestionario'],
            titulo=data['titulo'],
            descripcion=data['descripcion'],
            es_publico=data['es_publico'],
            #id_usuario=data['id_usuario'],
            id_categoria=data['id_categoria'],
            vigente=data['vigente']
        )
        return jsonify({"mensaje": "Cuestionario actualizado con éxito"}), 201
    except KeyError:
        return jsonify({"error": "Faltan datos requeridos"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/cuestionarios/<int:id_cuestionario>", methods=['DELETE'])
def api_delete_cuestionario(id_cuestionario):
    try:
        # Llama a la función del controlador para desactivar
        exito = cc.desactivar(id_cuestionario)

        if exito:
            return jsonify({"mensaje": "Cuestionario desactivado correctamente"}), 200
        else:
            return jsonify({"error": "Cuestionario no encontrado"}), 404

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/cuestionarios/guardar-completo", methods=['POST'])
def api_guardar_cuestionario_completo():
    data = request.json # Aquí recibes el objeto cuestionarioData completo

    # Llama a una nueva función del controlador que se encargue de todo
    # Esta función debe usar una "transacción" para asegurar que todo se guarde correctamente
    id_cuestionario_guardado = cc.guardar_o_actualizar_completo(data)

    return jsonify({"mensaje": "Cuestionario guardado con éxito", "id_cuestionario": id_cuestionario_guardado})

# TEST api Importar Excel
@app.route("/api/cuestionarios/guardar-basico", methods=['POST'])
def api_guardar_cuestionario_basico():
    """
    Crea un nuevo registro de CUESTIONARIO solo con los datos básicos
    y devuelve el ID generado. Usado antes de importar preguntas a un cuestionario nuevo.
    """
    if 'user_id' not in session:
        return jsonify({"success": False, "error": "No autorizado. Debes iniciar sesión."}), 401

    try:
        data = request.json
        if not data:
            return jsonify({"success": False, "error": "No se recibieron datos."}), 400

        # Obtener datos del JSON
        titulo = data.get('titulo', '').strip()
        id_categoria = data.get('id_categoria') # Dejar que el controlador valide si es None o inválido
        descripcion = data.get('descripcion', '').strip()
        es_publico = data.get('es_publico', False) # Default a False si no viene

        # Validar campos obligatorios
        if not titulo:
            return jsonify({"success": False, "error": "El título es obligatorio."}), 400
        if id_categoria is None: # Chequeo explícito por None
             return jsonify({"success": False, "error": "La categoría es obligatoria."}), 400
        try:
             id_categoria = int(id_categoria) # Asegurar que sea entero
        except (ValueError, TypeError):
             return jsonify({"success": False, "error": "La categoría seleccionada es inválida."}), 400


        # Datos fijos y de sesión
        id_usuario = session['user_id']
        vigente = 1 # Asumir que siempre se crea vigente
        fecha_creacion = datetime.now() # Fecha actual del servidor

        # --- Llamada al Controlador ---
        # Asumimos que tienes (o crearás) una función en 'controladores/cuestionarios.py'
        # llamada 'crear_solo_cuestionario' que recibe estos datos, inserta en la tabla
        # CUESTIONARIO y devuelve el ID insertado o None/lanza excepción si falla.
        # Debe manejar su propia conexión y commit/rollback.

        # Ejemplo de cómo podría ser la llamada (ajusta según tu controlador real):
        try:
            nuevo_id = cc.crear_solo_cuestionario(
                titulo=titulo,
                descripcion=descripcion,
                es_publico=bool(es_publico), # Asegurar booleano
                fecha_hora_creacion=fecha_creacion,
                id_usuario=id_usuario,
                id_categoria=id_categoria,
                vigente=vigente
            )

            if nuevo_id:
                return jsonify({"success": True, "id_cuestionario": nuevo_id}), 201 # 201 Created
            else:
                # Si el controlador devuelve None sin lanzar excepción (raro pero posible)
                return jsonify({"success": False, "error": "No se pudo crear el cuestionario en la base de datos."}), 500

        except pymysql.Error as db_err: # Capturar errores específicos de la BD
            print(f"Error DB al crear cuestionario básico: {db_err}")
            # Puedes ser más específico con el mensaje si quieres
            return jsonify({"success": False, "error": f"Error en base de datos: {db_err}"}), 500
        except Exception as ctrl_err: # Capturar otros errores del controlador
            print(f"Error en controlador al crear cuestionario básico: {ctrl_err}")
            return jsonify({"success": False, "error": f"Error al procesar la solicitud: {ctrl_err}"}), 500

    except Exception as e:
        # Captura errores generales (ej. JSON malformado)
        print(f"Error general en /guardar-basico: {e}")
        return jsonify({"success": False, "error": "Error interno del servidor."}), 500


# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# PREGUNTAS
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------

@app.route("/api_obtener_preguntas", methods=['GET'])
@jwt_required()
def fn_api_obtener_preguntas():
    try:
        cuestionarios = cpo.obtener_preguntas_todos()
        return jsonify({"data":cuestionarios}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500 

@app.route("/api_obtener_pregunta_por_id/<int:id_pregunta>", methods=['GET'])
@jwt_required()
def fn_api_obtener_pregunta_por_id(id_pregunta):
    try:
        pregunta = cpo.obtener_pregunta_por_id(id_pregunta)
        return jsonify({"data": pregunta}) if pregunta else (jsonify({"error": "Pregunta no encontrada"}), 404)
    except Exception as e:
        return jsonify({"error": str(e)}), 500 

@app.route("/api_registrar_pregunta", methods=['POST'])
@jwt_required()
def fn_api_registrar_pregunta():
    try:
        data = request.json
        
        id_cuestionario=data['id_cuestionario'],
        pregunta=data['pregunta'],
        num_pregunta=data['num_pregunta'],
        puntaje_base=data['puntaje_base'],
        tiempo=data.get('tiempo'),
        adjunto=data.get('adjunto')

        if tiempo < 10 or tiempo > 100:
            return jsonify({"error": "Tiempo de pregunta no inválido"}), 404
        if puntaje_base > 1000 or puntaje_base < 0:
            return jsonify({"error": "Puntaje de pregunta no inválido"}), 404
        
        
        cpo.crear_pregunta(
            id_cuestionario,
            pregunta,
            num_pregunta,
            puntaje_base,
            tiempo,
            adjunto
        )
        return jsonify({"mensaje": "Pregunta creada con éxito"}), 201
    except KeyError:
        return jsonify({"error": "Faltan datos requeridos"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api_actualizar_pregunta", methods=['PUT'])
@jwt_required()
def fn_api_actualizar_pregunta():
    try:
        data = request.json
        cpo.actualizar_pregunta(
            data['id_pregunta'],
            data['pregunta'], 
            data['puntaje_base']
        )
        return jsonify({"mensaje": "Pregunta actualizada"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500 

@app.route("/api_eliminar_pregunta", methods=['POST','DELETE'])
@jwt_required()
def fn_api_eliminar_pregunta():
    try:
        data = request.json
        # cpo.eliminar_pregunta(data['id_pregunta'])
        cpo.eliminar_pregunta_logica(data['id_pregunta'])
        return jsonify({"mensaje": "Pregunta eliminada logicamente"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------


@app.route("/api/preguntas/<int:id_pregunta>/imagen", methods=['POST'])
def api_upload_pregunta_imagen(id_pregunta):
    if 'imagen' not in request.files:
        return jsonify({"error": "No se encontró el archivo"}), 400

    file = request.files['imagen']
    if file.filename == '':
        return jsonify({"error": "No se seleccionó ningún archivo"}), 400

    if file and allowed_file(file.filename):
        # 1. Obtenemos el id_cuestionario para crear la carpeta correcta
        pregunta = cpo.obtener_pregunta_por_id(id_pregunta)
        if not pregunta:
            return jsonify({"error": "La pregunta no existe"}), 404
        id_cuestionario = pregunta['id_cuestionario']

        # 2. Creamos un nombre de archivo seguro y único
        ext = file.filename.rsplit('.', 1)[1].lower()
        timestamp = int(time.time())
        filename = secure_filename(f"pregunta_{id_pregunta}_{timestamp}.{ext}")

        # 3. Creamos la ruta y la carpeta si no existe
        upload_folder = os.path.join(app.root_path, 'static', 'img', 'cuestionarios', f'c_{id_cuestionario}')
        os.makedirs(upload_folder, exist_ok=True)

        # 4. Guardamos el archivo en el servidor
        file.save(os.path.join(upload_folder, filename))

        # 5. Guardamos la ruta relativa en la base de datos
        ruta_para_bd = f"img/cuestionarios/c_{id_cuestionario}/{filename}"
        cpo.actualizar_ruta_imagen(id_pregunta, ruta_para_bd)

        return jsonify({"mensaje": "Imagen subida con éxito", "ruta_img": ruta_para_bd}), 200

    return jsonify({"error": "Tipo de archivo no permitido"}), 400

@app.route("/api/preguntas/<int:id_pregunta>/imagen", methods=['DELETE'])
def api_delete_pregunta_imagen(id_pregunta):
    exito = cpo.quitar_imagen_pregunta(id_pregunta)
    if exito:
        return jsonify({"mensaje": "Imagen eliminada con éxito"}), 200
    else:
        return jsonify({"error": "No se pudo eliminar la imagen"}), 500


# --- PREGUNTAS ---
@app.route("/api/obtener_preguntas", methods=['GET'])
def api_obtener_preguntas():
    try:
        cuestionarios = cpo.obtener_preguntas_todos()
        return jsonify({"data":cuestionarios}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500 

@app.route("/api/cuestionarios/<int:id_cuestionario>/preguntas", methods=['GET'])
def api_get_preguntas(id_cuestionario):
    preguntas = cpo.obtener_preguntas_por_cuestionario(id_cuestionario)
    return jsonify({"data": preguntas}), 200


@app.route("/api/cuestionarios/<int:id_cuestionario>/preguntas", methods=['POST'])
def api_create_pregunta(id_cuestionario):
    try:
        data = request.json
        cpo.crear_pregunta(
            id_cuestionario=id_cuestionario,
            pregunta=data['pregunta'],
            num_pregunta=data['num_pregunta'],
            puntaje_base=data['puntaje_base'],
            adjunto=data.get('adjunto')
        )
        return jsonify({"mensaje": "Pregunta creada con éxito"}), 201
    except KeyError:
        return jsonify({"error": "Faltan datos requeridos"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/preguntas/<int:id_pregunta>", methods=['GET'])
def api_get_pregunta_por_id(id_pregunta):
    pregunta = cpo.obtener_pregunta_por_id(id_pregunta)
    return jsonify({"data": pregunta}) if pregunta else (jsonify({"error": "Pregunta no encontrada"}), 404)


@app.route("/api/preguntas/<int:id_pregunta>", methods=['PUT'])
def api_update_pregunta(id_pregunta):
    data = request.json
    cpo.actualizar_pregunta(id_pregunta, data['pregunta'], data['puntaje_base'])
    return jsonify({"mensaje": "Pregunta actualizada"})


@app.route("/api/preguntas/<int:id_pregunta>", methods=['DELETE'])
def api_delete_pregunta(id_pregunta):
    cpo.eliminar_pregunta(id_pregunta)
    return jsonify({"mensaje": "Pregunta eliminada"})



# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# OPCIONES
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------

@app.route("/api_obtener_opciones", methods=['GET'])
@jwt_required()
def fn_api_obtener_opciones():
    try:
        cuestionarios = cpo.obtener_opciones_todas()
        return jsonify({"data":cuestionarios}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500 

@app.route("/api_obtener_opcion_por_id/<int:id_opcion>", methods=['GET'])
@jwt_required()
def fn_api_get_opcion_por_id(id_opcion):
    try:
        opcion = cpo.obtener_opcion_por_id(id_opcion)
        return jsonify({"data": opcion}) if opcion else (jsonify({"error": "Opción no encontrada"}), 404)
    except Exception as e:
            return jsonify({"error": str(e)}), 500 

@app.route("/api_registar_opcion", methods=['POST'])
@jwt_required()
def fn_api_registrar_opcion():
    try:
        data = request.json
        cpo.crear_opcion(
            id_pregunta=data['id_pregunta'],
            opcion=data['opcion'],
            es_correcta_bool=data['es_correcta_bool']
            # descripcion=data.get('descripcion'),
            # adjunto=data.get('adjunto')
        )
        return jsonify({"mensaje": "Opción creada con éxito"}), 201
    except KeyError:
        return jsonify({"error": "Faltan datos requeridos"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api_actualizar_opcion", methods=['PUT'])
@jwt_required()
def fn_api_actualizar_opcion():
    try:
        data = request.json
        cpo.actualizar_opcion(data['id_opcion'], data['opcion'])
        return jsonify({"mensaje": "Opción actualizada"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api_eliminar_opcion", methods=['DELETE'])
@jwt_required()
def fn_api_delete_opcion():
    try:
        data = request.json
        # cpo.eliminar_opcion(data['id_opcion'])
        cpo.eliminar_opcion_logica(data['id_opcion'])
        return jsonify({"mensaje": "Opción eliminada"})
    except Exception as e:
            return jsonify({"error": str(e)}), 500


# ------------------------------------------------------------------------------
@app.route("/api/preguntas/<int:id_pregunta>/opciones", methods=['GET'])
def api_get_opciones(id_pregunta):
    opciones = cpo.obtener_opciones_por_pregunta(id_pregunta)
    return jsonify({"data": opciones}), 200

@app.route("/api/preguntas/<int:id_pregunta>/opciones", methods=['POST'])
def api_create_opcion(id_pregunta):
    try:
        data = request.json
        cpo.crear_opcion(
            id_pregunta=id_pregunta,
            opcion=data['opcion'],
            es_correcta_bool=data['es_correcta_bool'],
            descripcion=data.get('descripcion'),
            adjunto=data.get('adjunto')
        )
        return jsonify({"mensaje": "Opción creada con éxito"}), 201
    except KeyError:
        return jsonify({"error": "Faltan datos requeridos"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/opciones/<int:id_opcion>", methods=['GET'])
def api_get_opcion_por_id(id_opcion):
    opcion = cpo.obtener_opcion_por_id(id_opcion)
    return jsonify({"data": opcion}) if opcion else (jsonify({"error": "Opción no encontrada"}), 404)

@app.route("/api/opciones/<int:id_opcion>", methods=['PUT'])
def api_update_opcion(id_opcion):
    data = request.json
    cpo.actualizar_opcion(id_opcion, data['opcion'])
    return jsonify({"mensaje": "Opción actualizada"})

@app.route("/api/opciones/<int:id_opcion>", methods=['DELETE'])
def api_delete_opcion(id_opcion):
    cpo.eliminar_opcion(id_opcion)
    return jsonify({"mensaje": "Opción eliminada"})

# ------------------------------------------------------------------------------
# PARTIDAS (SESIONES DE JUEGO)

@app.route("/api/partidas", methods=['POST'])
def api_create_partida():
    # Lógica para crear una partida a partir de un id_cuestionario
    # Debería generar un PIN y devolverlo
    return jsonify({"mensaje": "Inicia una nueva partida y devuelve el PIN"}), 201

@app.route("/api/partidas/pin/<string:pin>", methods=['GET'])
def api_get_partida_by_pin(pin):
    # Lógica para que un jugador obtenga datos de la partida antes de unirse
    return jsonify({"mensaje": f"Devuelve info de la partida con PIN {pin}"})

@app.route("/api/partidas/<int:id_partida>/participantes", methods=['POST'])
def api_join_partida(id_partida):
    # Lógica para que un participante se una a una partida
    return jsonify({"mensaje": f"Un jugador se une a la partida {id_partida}"}), 201

@app.route("/api/partidas/<int:id_partida>/respuestas", methods=['POST'])
def api_submit_respuesta(id_partida):
    # Lógica para que un participante envíe su respuesta
    # Recibe: id_participante, id_opcion, tiempo_respuesta
    return jsonify({"mensaje": "Un jugador envía una respuesta"})

    #-------------------RECOMPENSAS----------------------------------

@app.route("/api/partidas/<int:id_partida>/finalizar", methods=['POST'])
def api_finalizar_partida(id_partida):
    """Finaliza una partida, otorga recompensas y muestra cambios en monedas."""
    import traceback
    import pymysql

    try:
        conexion = obtener_conexion()
        with conexion.cursor(pymysql.cursors.DictCursor) as cursor:
            # Verificar si la partida existe
            cursor.execute("SELECT estado FROM PARTIDA WHERE id_partida = %s", (id_partida,))
            partida = cursor.fetchone()
            if not partida:
                return jsonify({"error": f"La partida {id_partida} no existe"}), 404
            if partida["estado"] == "F":
                return jsonify({"mensaje": f"La partida {id_partida} ya estaba finalizada"}), 200

            # Obtener datos antes de otorgar recompensas
            cursor.execute("""
                SELECT p.id_usuario, u.nombre_completo, p.puntaje, u.monedas
                FROM PARTICIPANTE p
                JOIN USUARIO u ON p.id_usuario = u.id_usuario
                WHERE p.id_partida = %s
            """, (id_partida,))
            antes = cursor.fetchall()

        # Finalizar partida
        exito = c_part.finalizar_partida(id_partida)
        if not exito:
            return jsonify({"error": "No se pudo finalizar la partida"}), 500

        # Obtener datos después de otorgar recompensas
        conexion = obtener_conexion()
        with conexion.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("""
                SELECT p.id_usuario, u.nombre_completo, p.puntaje, u.monedas
                FROM PARTICIPANTE p
                JOIN USUARIO u ON p.id_usuario = u.id_usuario
                WHERE p.id_partida = %s
            """, (id_partida,))
            despues = cursor.fetchall()

        conexion.close()

        # Comparar resultados
        resumen = []
        for a in antes:
            d = next((x for x in despues if x["id_usuario"] == a["id_usuario"]), None)
            if d:
                resumen.append({
                    "id_usuario": a["id_usuario"],
                    "participante": a["nombre_completo"],
                    "puntaje": a["puntaje"],
                    "monedas_antes": a["monedas"],
                    "monedas_despues": d["monedas"],
                    "diferencia": d["monedas"] - a["monedas"]
                })

        return jsonify({
            "mensaje": f"Partida {id_partida} finalizada y recompensas otorgadas correctamente",
            "recompensas": resumen
        }), 200

    except Exception as e:
        return jsonify({
            "error": "Error interno al finalizar la partida",
            "detalle": str(e),
            "traza": traceback.format_exc()
        }), 500

# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# CRUD PARTIDA_JWS 
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------

def _serialize_partida_row(row): # Función para convertir los datetime a strings
    if not row:
        return None
    serializable = dict(row)
    for campo in ("fecha_hora_inicio", "fecha_hora_fin"):
        valor = serializable.get(campo)
        if isinstance(valor, datetime):
            serializable[campo] = valor.isoformat()
    return serializable

@app.route("/api_registrarpartida", methods=["POST"])
@jwt_required()
def api_registrarpartida():
    """
    Crea un registro completo en la tabla PARTIDA.
    """
    rpta = {"code": 0, "data": {}, "message": ""}
    conexion = None
    try:
        data = request.get_json(silent=True) or {}
        pin = data.get("pin")
        id_cuestionario = data.get("id_cuestionario")
        modalidad = data.get("modalidad")
        estado = data.get("estado")
        fecha_hora_inicio = data.get("fecha_hora_inicio")

        if not all([pin, id_cuestionario, modalidad, estado, fecha_hora_inicio]):
            rpta["message"] = "pin, id_cuestionario, modalidad, estado y fecha_hora_inicio son obligatorios."
            return jsonify(rpta), 400

        campos = [
            ("pin", pin),
            ("id_cuestionario", id_cuestionario),
            ("id_usuario_anfitrion", data.get("id_usuario_anfitrion")),
            ("modalidad", modalidad),
            ("estado", estado),
            ("fecha_hora_inicio", fecha_hora_inicio),
            ("fecha_hora_fin", data.get("fecha_hora_fin")),
            ("cant_grupos", data.get("cant_grupos")),
            ("recompensas_otorgadas", data.get("recompensas_otorgadas", 0)),
        ]

        columnas = ", ".join(col for col, val in campos if val is not None)
        valores = [val for _, val in campos if val is not None]
        placeholders = ", ".join(["%s"] * len(valores))

        conexion = obtener_conexion()
        with conexion.cursor() as cursor:
            cursor.execute(
                f"INSERT INTO PARTIDA ({columnas}) VALUES ({placeholders})",
                valores
            )
            conexion.commit()
            nuevo_id = cursor.lastrowid

        rpta["code"] = 1
        rpta["data"] = {"id_partida": nuevo_id}
        rpta["message"] = "Partida creada correctamente"
        return jsonify(rpta), 201
    except Exception as e:
        if conexion:
            conexion.rollback()
        rpta["message"] = f"No se pudo crear la partida: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()

@app.route("/api_obtenerpartidas", methods=["GET"])
@jwt_required()
def api_obtenerpartidas():
    rpta = {"code": 0, "data": {}, "message": ""}
    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("SELECT * FROM PARTIDA ORDER BY id_partida DESC")
            partidas = [_serialize_partida_row(row) for row in cursor.fetchall()]

        rpta["code"] = 1
        rpta["data"] = partidas
        rpta["message"] = "Partidas obtenidas correctamente"
        return jsonify(rpta)
    except Exception as e:
        rpta["message"] = f"No se pudieron obtener las partidas: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()
 
@app.route("/api_obtenerpartidaporid/<int:id_partida>", methods=["GET"])
@jwt_required()
def api_obtenerpartidaporid(id_partida):
    rpta = {"code": 0, "data": {}, "message": ""}
    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("SELECT * FROM PARTIDA WHERE id_partida = %s", (id_partida,))
            partida = cursor.fetchone()

        if not partida:
            rpta["message"] = "Partida no encontrada"
            return jsonify(rpta), 404

        rpta["code"] = 1
        rpta["data"] = _serialize_partida_row(partida)
        rpta["message"] = "Partida obtenida correctamente"
        return jsonify(rpta)
    except Exception as e:
        rpta["message"] = f"No se pudo obtener la partida: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()

@app.route("/api_actualizarpartida/<int:id_partida>", methods=["PUT"])
@jwt_required()
def api_actualizarpartida(id_partida):
    rpta = {"code": 0, "data": {}, "message": ""}
    data = request.get_json(silent=True) or {}
    campos_validos = {
        "pin", "id_cuestionario", "id_usuario_anfitrion", "modalidad",
        "estado", "fecha_hora_inicio", "fecha_hora_fin", "cant_grupos",
        "recompensas_otorgadas"
    }
    sets = []
    valores = []
    for campo in campos_validos:
        if campo in data:
            sets.append(f"{campo} = %s")
            valores.append(data[campo])

    if not sets:
        rpta["message"] = "No se enviaron campos a actualizar."
        return jsonify(rpta), 400

    valores.append(id_partida)
    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor() as cursor:
            cursor.execute(
                f"UPDATE PARTIDA SET {', '.join(sets)} WHERE id_partida = %s",
                valores
            )
            conexion.commit()
            if cursor.rowcount == 0:
                rpta["message"] = "Partida no encontrada"
                return jsonify(rpta), 404

        rpta["code"] = 1
        rpta["message"] = "Partida actualizada correctamente"
        return jsonify(rpta)
    except Exception as e:
        if conexion:
            conexion.rollback()
        rpta["message"] = f"No se pudo actualizar la partida: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()

@app.route("/api_eliminarpartida/<int:id_partida>", methods=["DELETE"])
@jwt_required()
def api_eliminarpartida(id_partida):
    rpta = {"code": 0, "data": {}, "message": ""}
    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor() as cursor:
            cursor.execute("DELETE FROM PARTIDA WHERE id_partida = %s", (id_partida,))
            conexion.commit()
            if cursor.rowcount == 0:
                rpta["message"] = "Partida no encontrada"
                return jsonify(rpta), 404

        rpta["code"] = 1
        rpta["message"] = "Partida eliminada correctamente"
        return jsonify(rpta)
    except Exception as e:
        if conexion:
            conexion.rollback()
        rpta["message"] = f"No se pudo eliminar la partida: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()


# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# SKINS - APIS CRUD
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------

@app.route("/api_registrarskin", methods=["POST"])
@jwt_required()
def api_registrarskin():
    """
    Crea un registro completo en la tabla SKINS.
    Campos: ruta (obligatorio), tipo (marco/fondo), precio, vigente
    """
    rpta = {"code": 0, "data": {}, "message": ""}
    conexion = None
    try:
        data = request.get_json(silent=True) or {}
        ruta = data.get("ruta")
        tipo = data.get("tipo", "marco")  # Por defecto 'marco'
        precio = data.get("precio", 100)

        if not ruta:
            rpta["message"] = "ruta es obligatorio."
            return jsonify(rpta), 400

        # Validar tipo
        if tipo not in ['marco', 'fondo']:
            rpta["message"] = "tipo debe ser 'marco' o 'fondo'."
            return jsonify(rpta), 400

        campos = [
            ("ruta", ruta),
            ("tipo", tipo),
            ("precio", precio),
            ("vigente", data.get("vigente", 1)),
        ]

        columnas = ", ".join(col for col, val in campos if val is not None)
        valores = [val for _, val in campos if val is not None]
        placeholders = ", ".join(["%s"] * len(valores))

        conexion = obtener_conexion()
        with conexion.cursor() as cursor:
            cursor.execute(
                f"INSERT INTO SKINS ({columnas}) VALUES ({placeholders})",
                valores
            )
            conexion.commit()
            nuevo_id = cursor.lastrowid

        rpta["code"] = 1
        rpta["data"] = {"id_skin": nuevo_id}
        rpta["message"] = "Skin creada correctamente"
        return jsonify(rpta), 201
    except Exception as e:
        if conexion:
            conexion.rollback()
        rpta["message"] = f"No se pudo crear la skin: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()

@app.route("/api_obtenerskins", methods=["GET"])
@jwt_required()
def api_obtenerskins():
    rpta = {"code": 0, "data": {}, "message": ""}
    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("SELECT * FROM SKINS ORDER BY id_skin DESC")
            skins = cursor.fetchall()

        rpta["code"] = 1
        rpta["data"] = skins
        rpta["message"] = "Skins obtenidas correctamente"
        return jsonify(rpta)
    except Exception as e:
        rpta["message"] = f"No se pudieron obtener las skins: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()

@app.route("/api_obtenerskinporid/<int:id_skin>", methods=["GET"])
@jwt_required()
def api_obtenerskinporid(id_skin):
    rpta = {"code": 0, "data": {}, "message": ""}
    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("SELECT * FROM SKINS WHERE id_skin = %s", (id_skin,))
            skin = cursor.fetchone()

        if not skin:
            rpta["message"] = "Skin no encontrada"
            return jsonify(rpta), 404

        rpta["code"] = 1
        rpta["data"] = skin
        rpta["message"] = "Skin obtenida correctamente"
        return jsonify(rpta)
    except Exception as e:
        rpta["message"] = f"No se pudo obtener la skin: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()

@app.route("/api_actualizarskin/<int:id_skin>", methods=["PUT"])
@jwt_required()
def api_actualizarskin(id_skin):
    rpta = {"code": 0, "data": {}, "message": ""}
    data = request.get_json(silent=True) or {}
    campos_validos = {
        "ruta", "tipo", "precio", "vigente"
    }
    sets = []
    valores = []
    for campo in campos_validos:
        if campo in data:
            # Validar tipo si se está actualizando
            if campo == "tipo" and data[campo] not in ['marco', 'fondo']:
                rpta["message"] = "tipo debe ser 'marco' o 'fondo'."
                return jsonify(rpta), 400
            sets.append(f"{campo} = %s")
            valores.append(data[campo])

    if not sets:
        rpta["message"] = "No se enviaron campos a actualizar."
        return jsonify(rpta), 400

    valores.append(id_skin)
    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor() as cursor:
            cursor.execute(
                f"UPDATE SKINS SET {', '.join(sets)} WHERE id_skin = %s",
                valores
            )
            conexion.commit()
            if cursor.rowcount == 0:
                rpta["message"] = "Skin no encontrada"
                return jsonify(rpta), 404

        rpta["code"] = 1
        rpta["message"] = "Skin actualizada correctamente"
        return jsonify(rpta)
    except Exception as e:
        if conexion:
            conexion.rollback()
        rpta["message"] = f"No se pudo actualizar la skin: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()

@app.route("/api_eliminarskin/<int:id_skin>", methods=["DELETE"])
@jwt_required()
def api_eliminarskin(id_skin):
    rpta = {"code": 0, "data": {}, "message": ""}
    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor() as cursor:
            cursor.execute("DELETE FROM SKINS WHERE id_skin = %s", (id_skin,))
            conexion.commit()
            if cursor.rowcount == 0:
                rpta["message"] = "Skin no encontrada"
                return jsonify(rpta), 404

        rpta["code"] = 1
        rpta["message"] = "Skin eliminada correctamente"
        return jsonify(rpta)
    except Exception as e:
        if conexion:
            conexion.rollback()
        rpta["message"] = f"No se pudo eliminar la skin: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()


# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# INVENTARIO SKINS - APIS CRUD
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------

@app.route("/api_registrarinventario", methods=["POST"])
@jwt_required()
def api_registrarinventario():
    rpta = {"code": 0, "data": {}, "message": ""}
    data = request.get_json(silent=True) or {}
    id_usuario = data.get("id_usuario")
    id_skin = data.get("id_skin")

    if not id_usuario or not id_skin:
        rpta["message"] = "id_usuario e id_skin son obligatorios."
        return jsonify(rpta), 400

    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor() as cursor:
            # Verificar si ya existe
            cursor.execute("""
                SELECT 1 FROM INVENTARIO_SKINS WHERE id_usuario = %s AND id_skin = %s
            """, (id_usuario, id_skin))
            if cursor.fetchone():
                rpta["message"] = "El registro ya existe en el inventario."
                return jsonify(rpta), 400

            cursor.execute("""
                INSERT INTO INVENTARIO_SKINS (id_usuario, id_skin)
                VALUES (%s, %s)
            """, (id_usuario, id_skin))
            conexion.commit()

        rpta["code"] = 1
        rpta["message"] = "Inventario registrado correctamente."
        return jsonify(rpta), 201
    except Exception as e:
        if conexion:
            conexion.rollback()
        rpta["message"] = f"No se pudo registrar el inventario: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()

@app.route("/api_actualizarinventario/<int:id_usuario>/<int:id_skin>", methods=["PUT"])
@jwt_required()
def api_actualizarinventario(id_usuario, id_skin):
    rpta = {"code": 0, "data": {}, "message": ""}
    data = request.get_json(silent=True) or {}
    nuevo_id_skin = data.get("id_skin")

    if not nuevo_id_skin:
        rpta["message"] = "Debe enviar el nuevo id_skin para actualizar."
        return jsonify(rpta), 400

    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor() as cursor:
            cursor.execute("""
                UPDATE INVENTARIO_SKINS
                SET id_skin = %s
                WHERE id_usuario = %s AND id_skin = %s
            """, (nuevo_id_skin, id_usuario, id_skin))
            conexion.commit()

            if cursor.rowcount == 0:
                rpta["message"] = "No se encontró el registro para actualizar."
                return jsonify(rpta), 404

        rpta["code"] = 1
        rpta["message"] = "Inventario actualizado correctamente."
        return jsonify(rpta)
    except Exception as e:
        if conexion:
            conexion.rollback()
        rpta["message"] = f"No se pudo actualizar el inventario: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()

@app.route("/api_eliminarinventario/<int:id_usuario>/<int:id_skin>", methods=["DELETE"])
@jwt_required()
def api_eliminarinventario(id_usuario, id_skin):
    rpta = {"code": 0, "data": {}, "message": ""}
    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor() as cursor:
            cursor.execute("""
                DELETE FROM INVENTARIO_SKINS
                WHERE id_usuario = %s AND id_skin = %s
            """, (id_usuario, id_skin))
            conexion.commit()

            if cursor.rowcount == 0:
                rpta["message"] = "No se encontró el registro a eliminar."
                return jsonify(rpta), 404

        rpta["code"] = 1
        rpta["message"] = "Inventario eliminado correctamente."
        return jsonify(rpta)
    except Exception as e:
        if conexion:
            conexion.rollback()
        rpta["message"] = f"No se pudo eliminar el inventario: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()

@app.route("/api_obtenerinventarios", methods=["GET"])
@jwt_required()
def api_obtenerinventarios():
    rpta = {"code": 0, "data": {}, "message": ""}
    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("""
                SELECT i.id_usuario, i.id_skin, u.nombre AS nombre_usuario, s.ruta AS ruta_skin
                FROM INVENTARIO_SKINS i
                INNER JOIN USUARIO u ON u.id_usuario = i.id_usuario
                INNER JOIN SKINS s ON s.id_skin = i.id_skin
                ORDER BY i.id_usuario
            """)
            registros = cursor.fetchall()

        rpta["code"] = 1
        rpta["data"] = registros
        rpta["message"] = "Inventarios obtenidos correctamente."
        return jsonify(rpta)
    except Exception as e:
        rpta["message"] = f"No se pudieron obtener los inventarios: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()

@app.route("/api_obtenerinventarioporusuario/<int:id_usuario>", methods=["GET"])
@jwt_required()
def api_obtenerinventarioporusuario(id_usuario):
    rpta = {"code": 0, "data": {}, "message": ""}
    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("""
                SELECT s.id_skin, s.ruta, s.tipo, s.precio, s.vigente
                FROM INVENTARIO_SKINS i
                INNER JOIN SKINS s ON s.id_skin = i.id_skin
                WHERE i.id_usuario = %s
            """, (id_usuario,))
            skins = cursor.fetchall()

        rpta["code"] = 1
        rpta["data"] = skins
        rpta["message"] = "Inventario del usuario obtenido correctamente."
        return jsonify(rpta)
    except Exception as e:
        rpta["message"] = f"No se pudo obtener el inventario: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()


# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# CATEGORIAS - APIS CRUD
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------

@app.route("/api_registrarcategoria", methods=["POST"])
@jwt_required()
def api_registrarcategoria():
    """
    Crea un registro completo en la tabla CATEGORIA.
    """
    rpta = {"code": 0, "data": {}, "message": ""}
    conexion = None
    try:
        data = request.get_json(silent=True) or {}
        categoria = data.get("categoria")

        if not categoria:
            rpta["message"] = "categoria es obligatorio."
            return jsonify(rpta), 400

        conexion = obtener_conexion()
        with conexion.cursor() as cursor:
            cursor.execute(
                "INSERT INTO CATEGORIA (categoria) VALUES (%s)",
                (categoria,)
            )
            conexion.commit()
            nuevo_id = cursor.lastrowid

        rpta["code"] = 1
        rpta["data"] = {"id_categoria": nuevo_id}
        rpta["message"] = "Categoría creada correctamente"
        return jsonify(rpta), 201
    except Exception as e:
        if conexion:
            conexion.rollback()
        rpta["message"] = f"No se pudo crear la categoría: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()


@app.route("/api_obtenercategorias", methods=["GET"])
@jwt_required()
def api_obtenercategorias():
    rpta = {"code": 0, "data": {}, "message": ""}
    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("SELECT * FROM CATEGORIA ORDER BY categoria ASC")
            categorias = cursor.fetchall()

        rpta["code"] = 1
        rpta["data"] = categorias
        rpta["message"] = "Categorías obtenidas correctamente"
        return jsonify(rpta)
    except Exception as e:
        rpta["message"] = f"No se pudieron obtener las categorías: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()


@app.route("/api_obtenercategoriaporid/<int:id_categoria>", methods=["GET"])
@jwt_required()
def api_obtenercategoriaporid(id_categoria):
    rpta = {"code": 0, "data": {}, "message": ""}
    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("SELECT * FROM CATEGORIA WHERE id_categoria = %s", (id_categoria,))
            categoria = cursor.fetchone()

        if not categoria:
            rpta["message"] = "Categoría no encontrada"
            return jsonify(rpta), 404

        rpta["code"] = 1
        rpta["data"] = categoria
        rpta["message"] = "Categoría obtenida correctamente"
        return jsonify(rpta)
    except Exception as e:
        rpta["message"] = f"No se pudo obtener la categoría: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()


@app.route("/api_actualizarcategoria/<int:id_categoria>", methods=["PUT"])
@jwt_required()
def api_actualizarcategoria(id_categoria):
    rpta = {"code": 0, "data": {}, "message": ""}
    data = request.get_json(silent=True) or {}
    categoria = data.get("categoria")

    if not categoria:
        rpta["message"] = "No se envió el campo categoria a actualizar."
        return jsonify(rpta), 400

    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor() as cursor:
            cursor.execute(
                "UPDATE CATEGORIA SET categoria = %s WHERE id_categoria = %s",
                (categoria, id_categoria)
            )
            conexion.commit()
            if cursor.rowcount == 0:
                rpta["message"] = "Categoría no encontrada"
                return jsonify(rpta), 404

        rpta["code"] = 1
        rpta["message"] = "Categoría actualizada correctamente"
        return jsonify(rpta)
    except Exception as e:
        if conexion:
            conexion.rollback()
        rpta["message"] = f"No se pudo actualizar la categoría: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()


@app.route("/api_eliminarcategoria/<int:id_categoria>", methods=["DELETE"])
@jwt_required()
def api_eliminarcategoria(id_categoria):
    rpta = {"code": 0, "data": {}, "message": ""}
    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor() as cursor:
            cursor.execute("DELETE FROM CATEGORIA WHERE id_categoria = %s", (id_categoria,))
            conexion.commit()
            if cursor.rowcount == 0:
                rpta["message"] = "Categoría no encontrada"
                return jsonify(rpta), 404

        rpta["code"] = 1
        rpta["message"] = "Categoría eliminada correctamente"
        return jsonify(rpta)
    except Exception as e:
        if conexion:
            conexion.rollback()
        rpta["message"] = f"No se pudo eliminar la categoría: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()


# ---------------------------------------------------------------------------------
# ---------------------------
# AJAX GAME (sin sockets)
# ---------------------------
# AJAX GAME (sin sockets)

@app.route('/api/game/host/init', methods=["POST"])
def api_game_host_init():
    data = request.get_json(force=True) or {}
    pin = (data.get('pin') or '').strip().upper()
    partida = _ensure_loaded(pin)
    if not partida:
        return jsonify({'error': 'PIN inválido o no disponible'}), 404
    return jsonify({'ok': True, 'lobby': get_lobby_state(pin), 'estado': partida['estado'], 'fase': partida['fase']})


@app.route('/api/lobby/state', methods=["GET"])
def api_lobby_state():
    pin = (request.args.get('pin') or '').strip().upper()
    if not pin:
        return jsonify({'error': 'pin requerido'}), 400
    if not _ensure_loaded(pin):
        return jsonify({'error': 'PIN inválido'}), 404
    return jsonify(get_lobby_state(pin))


@app.route('/api/player/join', methods=["POST"])
def api_player_join():
    data = request.get_json(force=True) or {}
    pin = (data.get('pin') or '').strip().upper()
    if not pin:
        return jsonify({'error': 'pin requerido'}), 400
    if not _ensure_loaded(pin):
        return jsonify({'error': 'PIN inválido'}), 404
    partida = join_player(pin)
    if not partida:
        return jsonify({'error': 'no se pudo unir'}), 400
    st = get_status(pin)
    return jsonify({'ok': True, 'estado': st.get('estado'), 'fase': st.get('fase'), 'lobby': get_lobby_state(pin)})


@app.route('/api/player/select-group', methods=["POST"])
def api_player_select_group():
    data = request.get_json(force=True) or {}
    pin = (data.get('pin') or '').strip().upper()
    nombre_grupo = (data.get('nombre_grupo') or '').strip()
    if not pin or not nombre_grupo:
        return jsonify({'error': 'pin y nombre_grupo requeridos'}), 400
    if not _ensure_loaded(pin):
        return jsonify({'error': 'PIN inválido'}), 404
    ok = select_group(pin, nombre_grupo)
    if not ok:
        return jsonify({'error': 'no se pudo seleccionar grupo'}), 400
    return jsonify({'ok': True, 'lobby': get_lobby_state(pin)})


@app.route('/api/player/leave', methods=["POST"])
def api_player_leave():
    data = request.get_json(force=True) or {}
    pin = (data.get('pin') or '').strip().upper()
    if not pin:
        return jsonify({'error': 'pin requerido'}), 400
    if not _ensure_loaded(pin):
        return jsonify({'error': 'PIN inválido'}), 404
    ok = remove_player(pin)
    if not ok:
        return jsonify({'error': 'no se pudo salir'}), 400
    return jsonify({'ok': True})


@app.route('/api/game/start', methods=["POST"])
def api_game_start():
    data = request.get_json(force=True) or {}
    pin = (data.get('pin') or '').strip().upper()
    if not pin:
        return jsonify({'error': 'pin requerido'}), 400
    if not _ensure_loaded(pin):
        return jsonify({'error': 'PIN inválido'}), 404
    started = start_game(pin)
    return jsonify({'started': bool(started)})


@app.route('/api/game/status', methods=["GET"])
def api_game_status():
    pin = (request.args.get('pin') or '').strip().upper()
    if not pin:
        return jsonify({'error': 'pin requerido'}), 400
    st = get_status(pin)
    if not st.get('existe'):
        return jsonify({'error': 'PIN inválido'}), 404
    resp = {'estado': st['estado'], 'fase': st['fase']}
    if st['estado'] == 'J':
        resp['url'] = url_for('pagina_juego', pin=pin)
    return jsonify(resp)


@app.route('/api/game/current', methods=["GET"])
def api_game_current():
    pin = (request.args.get('pin') or '').strip().upper()
    if not pin:
        return jsonify({'error': 'pin requerido'}), 400
    cur = get_current(pin)
    if not cur.get('existe'):
        return jsonify({'error': 'PIN inválido'}), 404
    return jsonify(cur)


@app.route('/api/game/answer', methods=["POST"])
def api_game_answer():
    data = request.get_json(force=True) or {}
    pin = (data.get('pin') or '').strip().upper()
    id_opcion = data.get('id_opcion')
    tiempo_restante = data.get('tiempo_restante', 0)
    if not pin or id_opcion is None:
        return jsonify({'error': 'pin e id_opcion requeridos'}), 400
    
    # Obtener info de la pregunta actual
    partida = partidas_en_juego.get(pin)
    respuesta_correcta = None
    es_correcta = False
    mi_puntaje = 0
    
    if not partida or partida['fase'] != 'question':
        return jsonify({'error': 'No hay pregunta activa'}), 400
    
    pregunta = partida['preguntas_data'][partida['pregunta_actual_index']]
    opciones = cpo.obtener_opciones_por_pregunta(pregunta['id_pregunta'])
    
    # Verificar si la opción seleccionada es correcta
    opcion_seleccionada = next((o for o in opciones if o['id_opcion'] == id_opcion), None)
    if opcion_seleccionada:
        es_correcta = opcion_seleccionada.get('es_correcta_bool', False)
    
    # Encontrar texto de la opción correcta
    for o in opciones:
        if o.get('es_correcta_bool'):
            respuesta_correcta = o['opcion']
            break
    
    # ENVIAR LA RESPUESTA (esto calcula y suma los puntos)
    ok = submit_answer(pin, id_opcion, tiempo_restante)
    
    # AHORA obtener el puntaje ACTUALIZADO
    if ok:
        nombre_usuario = session.get('nombre_usuario')
        if nombre_usuario and nombre_usuario in partida['participantes']:
            participante = partida['participantes'][nombre_usuario]
            if partida['modalidad_grupal']:
                nombre_grupo = participante.get('grupo')
                if nombre_grupo:
                    grupo = next((g for g in partida['grupos'] if g['nombre'] == nombre_grupo), None)
                    if grupo:
                        mi_puntaje = grupo['puntaje']
            else:
                mi_puntaje = participante['puntaje']
    
    return jsonify({
        'ok': bool(ok),
        'respuesta_correcta': respuesta_correcta,
        'es_correcta': es_correcta,
        'mi_puntaje': mi_puntaje,
        'nombre_quien_respondio': session.get('nombre_usuario', 'Compañero'),  # ⭐ Para modal grupal
        'texto_opcion': opcion_seleccionada.get('opcion', 'Opción desconocida') if opcion_seleccionada else 'Opción desconocida'  # ⭐ Para modal grupal
    })

#------------------CODIGO PARA ACCEDER AL CUESTIONARIO EN MODO VISUALIZACION------


@app.route("/api/validar-codigo-sesion", methods=['POST'])
def api_validar_codigo_sesion():
    """Valida el código (PIN) de una sesión y devuelve el id_cuestionario asociado"""
    try:
        data = request.get_json(force=True) or {}
        codigo = (data.get("codigo") or "").strip().upper()

        if not codigo or len(codigo) != 6:
            return jsonify({"error": "El código debe tener 6 caracteres"}), 400

        conexion = obtener_conexion()
        try:
            with conexion.cursor() as cursor:
                # Buscar partida con ese PIN que esté activa (estado 'E' o 'P')
                cursor.execute("""
                    SELECT p.id_partida, p.id_cuestionario, p.estado, c.titulo
                    FROM PARTIDA p
                    JOIN CUESTIONARIO c ON p.id_cuestionario = c.id_cuestionario
                    WHERE p.pin = %s AND p.estado IN ('E', 'P')
                    LIMIT 1
                """, (codigo,))

                partida = cursor.fetchone()

                if not partida:
                    return jsonify({"error": "Código de sesión inválido o expirado"}), 404

                return jsonify({
                    "ok": True,
                    "id_cuestionario": partida["id_cuestionario"],
                    "titulo": partida["titulo"],
                    "id_partida": partida["id_partida"]
                }), 200


        finally:
            conexion.close()

    except Exception as e:

        return jsonify({
            "error": "Error al validar el código",
            "detalle": str(e),
            "trace": traceback.format_exc()
        }), 500


#@app.route("/api/validar-codigo-sesion", methods=['POST'])
#def api_validar_codigo_sesion():
#    """Valida el código (PIN) de una sesión y devuelve el id_cuestionario asociado"""
#    try:
#        data = request.get_json(force=True) or {}
#        codigo = (data.get("codigo") or "").strip().upper()
#
#        if not codigo or len(codigo) != 6:
#            return jsonify({"error": "El código debe tener 6 caracteres"}), 400
#
#        conexion = obtener_conexion()
#        try:
#            with conexion.cursor() as cursor:
#                # Buscar partida con ese PIN (cualquier estado: E, P o F)
#                cursor.execute("""
#                    SELECT
#                        p.id_partida,
#                        p.id_cuestionario,
#                        p.estado,
#                        c.titulo,
#                        CASE
#                            WHEN p.estado = 'E' THEN 'En espera'
#                            WHEN p.estado = 'P' THEN 'En progreso'
#                           WHEN p.estado = 'F' THEN 'Finalizada'
#                        END as estado_texto
#                    FROM PARTIDA p
#                    JOIN CUESTIONARIO c ON p.id_cuestionario = c.id_cuestionario
#                    WHERE p.pin = %s
#                    LIMIT 1
#                """, (codigo,))
#
#                partida = cursor.fetchone()
#
#                if not partida:
#                    return jsonify({"error": "Código de sesión inválido"}), 404

#                # Validar según el estado
#                id_partida, id_cuestionario, estado, titulo, estado_texto = partida
#
#                # OPCIÓN A: Permitir todos los estados
#                return jsonify({
#                    "ok": True,
#                    "id_cuestionario": id_cuestionario,
#                    "titulo": titulo,
#                    "id_partida": id_partida,
#                    "estado": estado,
#                    "estado_texto": estado_texto
#                }), 200

                # OPCIÓN B: Solo permitir E y P (descomenta para usar)
                # if estado not in ('E', 'P'):
                #     return jsonify({
                #         "error": f"Esta sesión ya ha finalizado. Estado actual: {estado_texto}"
                #     }), 403

                # OPCIÓN C: Solo permitir E (descomenta para usar)
                # if estado != 'E':
                #     return jsonify({
                #         "error": f"Esta sesión ya comenzó o finalizó. Estado: {estado_texto}"
                #     }), 403

#        finally:
#            conexion.close()

    except Exception as e:
        return jsonify({"error": "Error al validar el código", "detalle": str(e)}), 500

# --- INICIO: NUEVO ENDPOINT PARA SOLICITAR RESTABLECIMIENTO ---

@app.route("/api/solicitar-restablecimiento", methods=['POST'])
def api_solicitar_restablecimiento():
    data = request.get_json()
    if not data or 'correo' not in data:
        return jsonify({"error": "Falta el campo 'correo'."}), 400

    correo = (data['correo'] or '').strip().lower()

    # Consulta "fuerte": solo trae si está vigente y verificado
    usuario_ok = ctrl.obtener_por_correo(correo)

    if not usuario_ok:
        # Consulta adicional para diferenciar si es "no verificado" o "no existe"
        usuario_any = ctrl.obtener_por_correo_sin_filtros(correo)
        if usuario_any and usuario_any.get('vigente') and not usuario_any.get('verificado'):
            return jsonify({"error": "Este correo no ha sido verificado. Por favor, verifica tu correo antes de restablecer la contraseña."}), 400
        else:
            # El correo no existe en la base de datos
            return jsonify({"error": "No existe una cuenta registrada con este correo electrónico."}), 404

    # Si llegó acá, está vigente y verificado → envía correo
    try:
        token = ctrl.generar_token_restablecimiento(usuario_ok['id_usuario'])
        url_de_restablecimiento = url_for('restablecer_con_token', token=token, _external=True)
        enviado, mensaje_email = email_sender.enviar_correo_restablecimiento(usuario_ok['correo'], url_de_restablecimiento)
        if not enviado:
            return jsonify({"error": f"No se pudo enviar el correo. Detalles: {mensaje_email}"}), 500
    except Exception as e:
        return jsonify({"error": f"Error interno del servidor: {str(e)}"}), 500

    return jsonify({"mensaje": "En breve recibirás un enlace para restablecer tu contraseña."}), 200
# --- FIN: NUEVO ENDPOINT ---

# ---------------------------------- AGREGADO POR PAME - Reportes
# --- GAME: submit answer ---
@app.route("/api/game/submit", methods=["POST"])
def api_game_submit():
    try:
        data = request.get_json() or {}
        pin = data.get("pin")
        id_opcion = data.get("id_opcion")
        tiempo_restante = data.get("tiempo_restante")  # segundos restantes (int)

        if not pin or not id_opcion:
            return jsonify({"ok": False, "msg": "Faltan pin o id_opcion"}), 400

        from ajax_game import submit_answer
        ok = submit_answer(pin, id_opcion, tiempo_restante)
        return jsonify({"ok": bool(ok)})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"ok": False, "msg": str(e)}), 500

# ---------- REPORTES JSON ----------
@app.route("/api/report/partida", methods=["GET"])
def api_report_partida():
    """
    Uso:
        /api/report/partida?pin=AB12CD
    ó  /api/report/partida?id_partida=123
    Devuelve JSON con header, resumen, preguntas, participantes.
    """
    try:
        pin = request.args.get("pin")
        id_partida = request.args.get("id_partida", type=int)

        if not pin and not id_partida:
            return jsonify({"ok": False, "msg": "Proporcione pin o id_partida"}), 400

        from controladores import controlador_partidas as c_part

        if pin:
            data = c_part.reporte_por_pin(pin)
        else:
            data = c_part.reporte_por_partida_id(id_partida)

        if not data:
            return jsonify({"ok": False, "msg": "Partida no encontrada"}), 404

        return jsonify({"ok": True, "data": data})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"ok": False, "msg": str(e)}), 500


# ---------- REPORTES CSV (descarga) ----------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Border, Side
from flask import send_file
import io

# @app.route("/api/report/partida/export", methods=["GET"])
# def api_report_partida_export():
#     """
#     Exporta Excel con múltiples hojas: Resumen, Detalle por participante, Detalle por pregunta
#     Uso: /api/report/partida/export?pin=AB12CD o /api/report/partida/export?id_partida=123
#     """
#     try:
#         pin = request.args.get("pin")
#         id_partida = request.args.get("id_partida", type=int)

#         from controladores import controlador_partidas as c_part

#         # Obtener datos completos del reporte
#         if pin:
#             data = c_part.reporte_por_pin(pin)
#         elif id_partida:
#             data = c_part.reporte_por_partida_id(id_partida)
#         else:
#             return jsonify({"ok": False, "msg": "Proporcione pin o id_partida"}), 400

#         target_partida_id = id_partida if id_partida else data['header']['id_partida']

#         # existing_report = None
#         # cx_lookup = obtener_conexion()
#         # try:
#         #     with cx_lookup.cursor() as c_lookup:
#         #         c_lookup.execute(
#         #             """
#         #             SELECT id_reporte, ruta
#         #                 FROM REPORTE
#         #                 WHERE id_partida=%s AND tipo='excel' AND subido_a='local'
#         #                 ORDER BY creado_en DESC
#         #                 LIMIT 1
#         #         """,
#         #             (target_partida_id,)
#         #         )
#         #         existing_report = c_lookup.fetchone()
#         # finally:
#         #     cx_lookup.close()

#         # Crear libro Excel
#         wb = Workbook()
#         wb.remove(wb.active)  # Eliminar hoja por defecto

#         def _set_dimensions(sheet):
#             max_col = sheet.max_column or 1
#             max_row = sheet.max_row or 1
#             for col_idx in range(1, max_col + 1):
#                 sheet.column_dimensions[get_column_letter(col_idx)].width = 20
#             for row_idx in range(1, max_row + 1):
#                 sheet.row_dimensions[row_idx].height = 20

#         # --- HOJA 1: RESUMEN (Ranking Final) ---
#         ws_resumen = wb.create_sheet("Resumen")
#         ws_resumen.append(["Posición", "Usuario/Grupo", "PuntajeTotal", "RespuestasCorrectas",
#                           "RespuestasIncorrectas", "%Acierto", "TiempoPromedioResp (s)"])

#         # Estilo para encabezados
#         for cell in ws_resumen[1]:
#             cell.font = Font(bold=True)
#             cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

#         # Llenar datos del ranking
#         for idx, part in enumerate(data['participantes'], start=1):
#             total_resp = int(part['correctas']) + int(part['incorrectas'])
#             pct_acierto = round((int(part['correctas']) / total_resp * 100), 2) if total_resp > 0 else 0
#             ws_resumen.append([
#                 idx,
#                 part['nombre'],
#                 int(part['puntaje_total']),
#                 int(part['correctas']),
#                 int(part['incorrectas']),
#                 pct_acierto,
#                 round(float(part['tiempo_prom_seg'] or 0), 2)
#             ])
#         _set_dimensions(ws_resumen)

#         # --- HOJA 2: DETALLE POR PARTICIPANTE ---
#         ws_detalle_part = wb.create_sheet("Detalle por participante")
#         ws_detalle_part.append(["Usuario", "Grupo(opc)", "Correcta",
#                                 "TiempoRestante", "PuntosOtorgados", "PuntajeAcumulado"])

#         for cell in ws_detalle_part[1]:
#             cell.font = Font(bold=True)
#             cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

#         # Obtener respuestas detalladas por participante
#         cx = obtener_conexion()
#         try:
#             with cx.cursor(pymysql.cursors.DictCursor) as c:
#                 c.execute("""
#                     SELECT pa.nombre, pa.id_grupo, rp.id_pregunta, rp.es_correcta,
#                           TIME_TO_SEC(rp.tiempo_respuesta) AS tiempo_seg, rp.puntaje,
#                           (SELECT SUM(rp2.puntaje)
#                             FROM RESPUESTA_PARTICIPANTE rp2
#                             WHERE rp2.id_participante = rp.id_participante
#                             AND rp2.id_respuesta <= rp.id_respuesta) AS puntaje_acum
#                     FROM RESPUESTA_PARTICIPANTE rp
#                     JOIN PARTICIPANTE pa ON pa.id_participante = rp.id_participante
#                     WHERE rp.id_partida=%s
#                     ORDER BY pa.nombre, rp.id_pregunta
#                 """, (target_partida_id,))
#                 rows_part = c.fetchall()

#                 for r in rows_part:
#                     grupo = r['id_grupo']
#                     ws_detalle_part.append([
#                         r['nombre'],
#                         grupo if grupo not in (None, "") else "No",
#                         "Sí" if int(r['es_correcta']) == 1 else "No",
#                         int(r['tiempo_seg'] or 0),
#                         int(r['puntaje']),
#                         int(r['puntaje_acum'] or 0)
#                     ])
#         finally:
#             cx.close()
#         _set_dimensions(ws_detalle_part)

#         # --- HOJA 3: DETALLE POR PREGUNTA ---
#         ws_detalle_preg = wb.create_sheet("Detalle por pregunta")
#         ws_detalle_preg.append(["#Pregunta", "Pregunta", "Opción", "Correcta",
#                               "RespuestasRecibidas", "Aciertos", "%Aciertos"])

#         for cell in ws_detalle_preg[1]:
#             cell.font = Font(bold=True)
#             cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

#         # Obtener detalle de opciones por pregunta
#         cx = obtener_conexion()
#         try:
#             with cx.cursor(pymysql.cursors.DictCursor) as c:
#                 c.execute("""
#                     SELECT rp.id_pregunta,
#                           MIN(rp.pregunta_texto) AS pregunta_texto,
#                           rp.opcion_texto,
#                           MAX(rp.es_correcta) AS es_correcta,
#                           COUNT(*) AS respuestas_recibidas,
#                           SUM(rp.es_correcta) AS aciertos,
#                           ROUND(AVG(TIME_TO_SEC(rp.tiempo_respuesta)), 2) AS tiempo_prom
#                     FROM RESPUESTA_PARTICIPANTE rp
#                     WHERE rp.id_partida=%s
#                     GROUP BY rp.id_pregunta, rp.opcion_texto
#                     ORDER BY rp.id_pregunta, rp.opcion_texto
#                 """, (target_partida_id,))
#                 rows_preg = c.fetchall()

#                 for r in rows_preg:
#                     pct_aciertos = round((int(r['aciertos']) / int(r['respuestas_recibidas']) * 100), 2) if r['respuestas_recibidas'] > 0 else 0
#                     ws_detalle_preg.append([
#                         r['id_pregunta'],
#                         r['pregunta_texto'],
#                         r['opcion_texto'],
#                         "Sí" if int(r['es_correcta']) == 1 else "No",
#                         int(r['respuestas_recibidas']),
#                         int(r['aciertos']),
#                         pct_aciertos,
#                     ])
#         finally:
#             cx.close()
#         _set_dimensions(ws_detalle_preg)

#         # timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
#         # filename = f"reporte_partida_{target_partida_id}_{timestamp}.xlsx"
#         #
#         # reports_dir = os.path.join(current_app.root_path, 'static', 'reportes', 'partidas')
#         # os.makedirs(reports_dir, exist_ok=True)
#         # abs_path = os.path.join(reports_dir, filename)
#         #
#         # wb.save(abs_path)
#         #
#         # rel_path = os.path.relpath(abs_path, os.path.join(current_app.root_path, 'static'))
#         # rel_path = rel_path.replace("\\", "/")
#         #
#         # if existing_report and existing_report.get('ruta'):
#         #     old_path = os.path.join(current_app.root_path, 'static', existing_report['ruta'].lstrip('/'))
#         #     if os.path.isfile(old_path):
#         #         try:
#         #             os.remove(old_path)
#         #         except OSError:
#         #             pass
#         #
#         # cx_db = obtener_conexion()
#         # try:
#         #     with cx_db.cursor() as cdb:
#         #         if existing_report:
#         #             cdb.execute(
#         #                 """
#         #                 UPDATE REPORTE
#         #                     SET ruta=%s,
#         #                         tipo='excel',
#         #                         subido_a='local',
#         #                         link_externo=NULL,
#         #                         creado_en=NOW()
#         #                 WHERE id_reporte=%s
#         #             """,
#         #                 (rel_path, existing_report['id_reporte'])
#         #             )
#         #         else:
#         #             cdb.execute(
#         #                 """
#         #                 INSERT INTO REPORTE (id_partida, ruta, tipo, subido_a)
#         #                 VALUES (%s, %s, %s, %s)
#         #             """,
#         #                 (target_partida_id, rel_path, 'excel', 'local')
#         #             )
#         #     cx_db.commit()
#         # finally:
#         #     cx_db.close()

#         output = io.BytesIO()
#         wb.save(output)
#         output.seek(0)
#         filename = f"reporte_partida_{target_partida_id}.xlsx"

#         return send_file(
#             output,
#             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#             as_attachment=True,
#             download_name=filename
#         )

#     except Exception as e:
#         traceback.print_exc()
#         return jsonify({"ok": False, "msg": str(e)}), 500


@app.route("/api/report/partida/export", methods=["GET"])
def api_report_partida_export():
    """
    Exporta Excel con múltiples hojas: Resumen, Detalle por participante, Detalle por pregunta
    Uso: /api/report/partida/export?pin=AB12CD o /api/report/partida/export?id_partida=123
    """
    try:
        pin = request.args.get("pin")
        id_partida = request.args.get("id_partida", type=int)

        from controladores import controlador_partidas as c_part
        
        # Obtener datos completos del reporte
        if pin:
            data = c_part.reporte_por_pin(pin)
        elif id_partida:
            data = c_part.reporte_por_partida_id(id_partida)
        else:
            return jsonify({"ok": False, "msg": "Proporcione pin o id_partida"}), 400

        target_partida_id = id_partida if id_partida else data['header']['id_partida']

        # existing_report = None
        # cx_lookup = obtener_conexion()
        # try:
        #     with cx_lookup.cursor() as c_lookup:
        #         c_lookup.execute(
        #             """
        #             SELECT id_reporte, ruta
        #                 FROM REPORTE
        #                 WHERE id_partida=%s AND tipo='excel' AND subido_a='local'
        #                 ORDER BY creado_en DESC
        #                 LIMIT 1
        #         """,
        #             (target_partida_id,)
        #         )
        #         existing_report = c_lookup.fetchone()
        # finally:
        #     cx_lookup.close()

        # Crear libro Excel
        wb = Workbook()
        wb.remove(wb.active)  # Eliminar hoja por defecto

        def _apply_sheet_style(sheet):
            max_col = sheet.max_column or 1
            max_row = sheet.max_row or 1
            center_alignment = Alignment(horizontal="center", vertical="center")
            for col_idx in range(1, max_col + 1):
                sheet.column_dimensions[get_column_letter(col_idx)].width = 25
            for row_idx in range(1, max_row + 1):
                sheet.row_dimensions[row_idx].height = 20
            for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    cell.alignment = center_alignment

        # --- HOJA 1: RESUMEN (Ranking Final) ---
        ws_resumen = wb.create_sheet("Resumen")

        header_data = data.get("header") or {}
        estado_map = {"F": "Finalizado", "E": "En espera", "P": "En progreso"}
        estado_texto = estado_map.get(header_data.get("estado"), header_data.get("estado") or "")

        def _to_text(value):
            if value is None:
                return ""
            if isinstance(value, datetime):
                return value.strftime("%Y-%m-%d %H:%M:%S")
            return str(value)

        ws_resumen["B1"] = "Datos de la PARTIDA:"
        ws_resumen["B1"].font = Font(bold=True)
        ws_resumen["B1"].alignment = Alignment(horizontal="left", vertical="center")

        info_rows = []
        if header_data:
            info_rows.extend([
                ("ID Partida", _to_text(header_data.get("id_partida"))),
                ("PIN", _to_text(header_data.get("pin"))),
                ("Cuestionario", _to_text(header_data.get("cuestionario_titulo"))),
                ("Estado", estado_texto),
            ])
            inicio = _to_text(header_data.get("fecha_hora_inicio"))
            if inicio:
                info_rows.append(("Fecha de inicio", inicio))
            fin = _to_text(header_data.get("fecha_hora_fin"))
            if fin:
                info_rows.append(("Fecha de finalizacion", fin))

        thin_side = Side(style="thin", color="000000")
        table_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        for row_offset, (label, value) in enumerate(info_rows, start=2):
            label_cell = ws_resumen.cell(row=row_offset, column=2, value=label)
            value_cell = ws_resumen.cell(row=row_offset, column=3, value=value)
            label_cell.font = Font(bold=True)
            label_cell.border = table_border
            value_cell.border = table_border

        resumen_data = data.get("resumen") or {}
        current_row = len(info_rows) + 3

        if resumen_data:
            resumen_title = ws_resumen.cell(row=current_row, column=2, value="Resumen global:")
            resumen_title.font = Font(bold=True)
            resumen_title.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1

            resumen_rows = [
                ("Participantes", _to_text(resumen_data.get("total_participantes"))),
                ("Preguntas distintas", _to_text(resumen_data.get("preguntas_distintas"))),
                ("Respuestas totales", _to_text(resumen_data.get("respuestas_totales"))),
                ("% Acierto global", f"{resumen_data.get('acierto_global')}%" if resumen_data.get("acierto_global") is not None else ""),
                ("Tiempo promedio (s)", _to_text(resumen_data.get("tiempo_promedio_seg"))),
            ]

            for label, value in resumen_rows:
                label_cell = ws_resumen.cell(row=current_row, column=2, value=label)
                value_cell = ws_resumen.cell(row=current_row, column=3, value=value)
                label_cell.font = Font(bold=True)
                label_cell.border = table_border
                value_cell.border = table_border
                current_row += 1

            current_row += 1  # Leave a blank row before the ranking table

        header_row_idx = current_row
        resumen_headers = [
            "Posicion",
            "Usuario",
            "Grupo",
            "PuntajeTotal",
            "Monedas",
            "RespuestasCorrectas",
            "RespuestasIncorrectas",
            "%Acierto",
            "TiempoPromedioResp (s)",
        ]

        for col_idx, value in enumerate(resumen_headers, start=1):
            cell = ws_resumen.cell(row=header_row_idx, column=col_idx, value=value)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFE400", end_color="FFE400", fill_type="solid")

        data_start_row = header_row_idx + 1
        for position, part in enumerate(data["participantes"], start=1):
            row = data_start_row + position - 1
            total_resp = int(part["correctas"]) + int(part["incorrectas"])
            pct_acierto = round((int(part["correctas"]) / total_resp * 100), 2) if total_resp > 0 else 0
            grupo_val = int(part.get("grupo") or 0)
            grupo_texto = grupo_val if grupo_val > 0 else "No"
            ws_resumen.cell(row=row, column=1, value=position)
            ws_resumen.cell(row=row, column=2, value=part["nombre"])
            ws_resumen.cell(row=row, column=3, value=grupo_texto)
            ws_resumen.cell(row=row, column=4, value=int(part["puntaje_total"]))
            ws_resumen.cell(row=row, column=5, value=int(part.get("monedas") or 0))
            ws_resumen.cell(row=row, column=6, value=int(part["correctas"]))
            ws_resumen.cell(row=row, column=7, value=int(part["incorrectas"]))
            ws_resumen.cell(row=row, column=8, value=pct_acierto)
            ws_resumen.cell(row=row, column=9, value=round(float(part["tiempo_prom_seg"] or 0), 2))


        _apply_sheet_style(ws_resumen)

        # --- HOJA 2: DETALLE POR PARTICIPANTE ---
        ws_detalle_part = wb.create_sheet("Detalle por participante")
        ws_detalle_part.append(["Usuario", "Grupo", "Correcta", 
                                "TiempoRestante", "PuntosOtorgados", "PuntajeAcumulado"])
        
        for cell in ws_detalle_part[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFE400", end_color="FFE400", fill_type="solid")

        # Obtener respuestas detalladas por participante
        cx = obtener_conexion()
        try:
            with cx.cursor(pymysql.cursors.DictCursor) as c:
                c.execute("""
                    SELECT pa.nombre, pa.id_grupo, rp.id_pregunta, rp.es_correcta,
                           pa.grupo AS numero_grupo,
                           TIME_TO_SEC(rp.tiempo_respuesta) AS tiempo_seg, rp.puntaje,
                           (SELECT SUM(rp2.puntaje) 
                            FROM RESPUESTA_PARTICIPANTE rp2 
                            WHERE rp2.id_participante = rp.id_participante 
                            AND rp2.id_respuesta <= rp.id_respuesta) AS puntaje_acum
                    FROM RESPUESTA_PARTICIPANTE rp
                    JOIN PARTICIPANTE pa ON pa.id_participante = rp.id_participante
                    WHERE rp.id_partida=%s
                    ORDER BY pa.nombre, rp.id_pregunta
                """, (target_partida_id,))
                rows_part = c.fetchall()
                
                for r in rows_part:
                    grupo_val = int(r.get('numero_grupo') or 0)
                    grupo_texto = grupo_val if grupo_val > 0 else "No"
                    ws_detalle_part.append([
                        r['nombre'],
                        grupo_texto,
                        "Sí" if int(r['es_correcta']) == 1 else "No",
                        int(r['tiempo_seg'] or 0),
                        int(r['puntaje']),
                        int(r['puntaje_acum'] or 0)
                    ])
        finally:
            cx.close()
        _apply_sheet_style(ws_detalle_part)

        # --- HOJA 3: DETALLE POR PREGUNTA ---
        ws_detalle_preg = wb.create_sheet("Detalle por pregunta")
        ws_detalle_preg.append(["#Pregunta", "Pregunta", "Opción", "Correcta", 
                               "RespuestasRecibidas", "Aciertos", "%Aciertos"])
        
        for cell in ws_detalle_preg[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFE400", end_color="FFE400", fill_type="solid")

        # Obtener detalle de opciones por pregunta
        cx = obtener_conexion()
        try:
            with cx.cursor(pymysql.cursors.DictCursor) as c:
                c.execute("""
                    SELECT rp.id_pregunta, 
                           MIN(rp.pregunta_texto) AS pregunta_texto,
                           rp.opcion_texto,
                           MAX(rp.es_correcta) AS es_correcta,
                           COUNT(*) AS respuestas_recibidas,
                           SUM(rp.es_correcta) AS aciertos,
                           ROUND(AVG(TIME_TO_SEC(rp.tiempo_respuesta)), 2) AS tiempo_prom
                    FROM RESPUESTA_PARTICIPANTE rp
                    WHERE rp.id_partida=%s
                    GROUP BY rp.id_pregunta, rp.opcion_texto
                    ORDER BY rp.id_pregunta, rp.opcion_texto
                """, (target_partida_id,))
                rows_preg = c.fetchall()
                
                for r in rows_preg:
                    pct_aciertos = round((int(r['aciertos']) / int(r['respuestas_recibidas']) * 100), 2) if r['respuestas_recibidas'] > 0 else 0
                    ws_detalle_preg.append([
                        r['id_pregunta'],
                        r['pregunta_texto'],
                        r['opcion_texto'],
                        "Sí" if int(r['es_correcta']) == 1 else "No",
                        int(r['respuestas_recibidas']),
                        int(r['aciertos']),
                        pct_aciertos,
                    ])
        finally:
            cx.close()
        _apply_sheet_style(ws_detalle_preg)

        # timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
        # filename = f"reporte_partida_{target_partida_id}_{timestamp}.xlsx"
        #
        # reports_dir = os.path.join(current_app.root_path, 'static', 'reportes', 'partidas')
        # os.makedirs(reports_dir, exist_ok=True)
        # abs_path = os.path.join(reports_dir, filename)
        #
        # wb.save(abs_path)
        #
        # rel_path = os.path.relpath(abs_path, os.path.join(current_app.root_path, 'static'))
        # rel_path = rel_path.replace("\\", "/")
        #
        # if existing_report and existing_report.get('ruta'):
        #     old_path = os.path.join(current_app.root_path, 'static', existing_report['ruta'].lstrip('/'))
        #     if os.path.isfile(old_path):
        #         try:
        #             os.remove(old_path)
        #         except OSError:
        #             pass
        #
        # cx_db = obtener_conexion()
        # try:
        #     with cx_db.cursor() as cdb:
        #         if existing_report:
        #             cdb.execute(
        #                 """
        #                 UPDATE REPORTE
        #                     SET ruta=%s,
        #                         tipo='excel',
        #                         subido_a='local',
        #                         link_externo=NULL,
        #                         creado_en=NOW()
        #                 WHERE id_reporte=%s
        #             """,
        #                 (rel_path, existing_report['id_reporte'])
        #             )
        #         else:
        #             cdb.execute(
        #                 """
        #                 INSERT INTO REPORTE (id_partida, ruta, tipo, subido_a)
        #                 VALUES (%s, %s, %s, %s)
        #             """,
        #                 (target_partida_id, rel_path, 'excel', 'local')
        #             )
        #     cx_db.commit()
        # finally:
        #     cx_db.close()

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"reporte_partida_{target_partida_id}.xlsx"

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({"ok": False, "msg": str(e)}), 500



@app.route("/api/report/partida/export-drive", methods=["GET"])
def api_report_partida_export_drive():
    """
    Genera Excel y lo sube a Google Drive en la carpeta 'DogeHoot Reportes' y lo comparte con un correo
    Uso: /api/report/partida/export-drive?pin=AB12CD&email=correo@ejemplo.com
    """
    try:
        pin = request.args.get("pin")
        id_partida = request.args.get("id_partida", type=int)
        email = request.args.get("email")
        
        # Validar que se proporcione el correo
        if not email:
            return jsonify({"ok": False, "msg": "Debe proporcionar un correo electrónico (parámetro 'email')"}), 400

        from controladores import controlador_partidas as c_part
        from controladores.google_drive_uploader import GoogleDriveUploader
        
        # Obtener datos completos del reporte
        if pin:
            data = c_part.reporte_por_pin(pin)
        elif id_partida:
            data = c_part.reporte_por_partida_id(id_partida)
        else:
            return jsonify({"ok": False, "msg": "Proporcione pin o id_partida"}), 400

        if not data:
            return jsonify({"ok": False, "msg": "Partida no encontrada"}), 404

        # Crear libro Excel (mismo código que export)
        wb = Workbook()
        wb.remove(wb.active)

        def _apply_sheet_style(sheet):
            max_col = sheet.max_column or 1
            max_row = sheet.max_row or 1
            center_alignment = Alignment(horizontal="center", vertical="center")
            for col_idx in range(1, max_col + 1):
                sheet.column_dimensions[get_column_letter(col_idx)].width = 25
            for row_idx in range(1, max_row + 1):
                sheet.row_dimensions[row_idx].height = 20
            for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    cell.alignment = center_alignment

        # --- HOJA 1: RESUMEN (Ranking Final) ---
        ws_resumen = wb.create_sheet("Resumen")

        header_data = data.get("header") or {}
        estado_map = {"F": "Finalizado", "E": "En espera", "P": "En progreso"}
        estado_texto = estado_map.get(header_data.get("estado"), header_data.get("estado") or "")

        def _to_text(value):
            if value is None:
                return ""
            if isinstance(value, datetime):
                return value.strftime("%Y-%m-%d %H:%M:%S")
            return str(value)

        ws_resumen["B1"] = "Datos de la PARTIDA:"
        ws_resumen["B1"].font = Font(bold=True)
        ws_resumen["B1"].alignment = Alignment(horizontal="left", vertical="center")

        info_rows = []
        if header_data:
            info_rows.extend([
                ("ID Partida", _to_text(header_data.get("id_partida"))),
                ("PIN", _to_text(header_data.get("pin"))),
                ("Cuestionario", _to_text(header_data.get("cuestionario_titulo"))),
                ("Estado", estado_texto),
            ])
            inicio = _to_text(header_data.get("fecha_hora_inicio"))
            if inicio:
                info_rows.append(("Fecha de inicio", inicio))
            fin = _to_text(header_data.get("fecha_hora_fin"))
            if fin:
                info_rows.append(("Fecha de finalizacion", fin))

        thin_side = Side(style="thin", color="000000")
        table_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        for row_offset, (label, value) in enumerate(info_rows, start=2):
            label_cell = ws_resumen.cell(row=row_offset, column=2, value=label)
            value_cell = ws_resumen.cell(row=row_offset, column=3, value=value)
            label_cell.font = Font(bold=True)
            label_cell.border = table_border
            value_cell.border = table_border

        header_row_idx = len(info_rows) + 3
        resumen_headers = [
            "Posicion",
            "Usuario",
            "Grupo",
            "PuntajeTotal",
            "Monedas",
            "RespuestasCorrectas",
            "RespuestasIncorrectas",
            "%Acierto",
            "TiempoPromedioResp (s)",
        ]

        for col_idx, value in enumerate(resumen_headers, start=1):
            cell = ws_resumen.cell(row=header_row_idx, column=col_idx, value=value)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFE400", end_color="FFE400", fill_type="solid")

        data_start_row = header_row_idx + 1
        for position, part in enumerate(data["participantes"], start=1):
            row = data_start_row + position - 1
            total_resp = int(part["correctas"]) + int(part["incorrectas"])
            pct_acierto = round((int(part["correctas"]) / total_resp * 100), 2) if total_resp > 0 else 0
            grupo_val = int(part.get("grupo") or 0)
            grupo_texto = grupo_val if grupo_val > 0 else "No"
            ws_resumen.cell(row=row, column=1, value=position)
            ws_resumen.cell(row=row, column=2, value=part["nombre"])
            ws_resumen.cell(row=row, column=3, value=grupo_texto)
            ws_resumen.cell(row=row, column=4, value=int(part["puntaje_total"]))
            ws_resumen.cell(row=row, column=5, value=int(part.get("monedas") or 0))
            ws_resumen.cell(row=row, column=6, value=int(part["correctas"]))
            ws_resumen.cell(row=row, column=7, value=int(part["incorrectas"]))
            ws_resumen.cell(row=row, column=8, value=pct_acierto)
            ws_resumen.cell(row=row, column=9, value=round(float(part["tiempo_prom_seg"] or 0), 2))


        _apply_sheet_style(ws_resumen)

        # --- HOJA 2: DETALLE POR PARTICIPANTE ---
        ws_detalle_part = wb.create_sheet("Detalle por participante")
        ws_detalle_part.append(["Usuario", "Grupo", "PreguntaN", "Correcta(0/1)", 
                                "TiempoRestante", "PuntosOtorgados", "PuntajeAcumulado"])
        
        for cell in ws_detalle_part[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFE400", end_color="FFE400", fill_type="solid")

        cx = obtener_conexion()
        try:
            with cx.cursor(pymysql.cursors.DictCursor) as c:
                c.execute("""
                    SELECT pa.nombre, pa.id_grupo, rp.id_pregunta, rp.es_correcta,
                           pa.grupo AS numero_grupo,
                           TIME_TO_SEC(rp.tiempo_respuesta) AS tiempo_seg, rp.puntaje,
                           (SELECT SUM(rp2.puntaje) 
                            FROM RESPUESTA_PARTICIPANTE rp2 
                            WHERE rp2.id_participante = rp.id_participante 
                            AND rp2.id_respuesta <= rp.id_respuesta) AS puntaje_acum
                    FROM RESPUESTA_PARTICIPANTE rp
                    JOIN PARTICIPANTE pa ON pa.id_participante = rp.id_participante
                    WHERE rp.id_partida=%s
                    ORDER BY pa.nombre, rp.id_pregunta
                """, (id_partida if id_partida else data['header']['id_partida'],))
                rows_part = c.fetchall()
                
                for r in rows_part:
                    ws_detalle_part.append([
                        r['nombre'],
                        (int(r.get('numero_grupo') or 0) or "No"),
                        r['id_pregunta'],
                        int(r['es_correcta']),
                        int(r['tiempo_seg'] or 0),
                        int(r['puntaje']),
                        int(r['puntaje_acum'] or 0)
                    ])
        finally:
            cx.close()
        _apply_sheet_style(ws_detalle_part)

        # --- HOJA 3: DETALLE POR PREGUNTA ---
        ws_detalle_preg = wb.create_sheet("Detalle por pregunta")
        ws_detalle_preg.append(["#Pregunta", "TextoPregunta", "Opción", "EsCorrecta(0/1)", 
                               "RespuestasRecibidas", "Aciertos", "%Aciertos", "TiempoPromedioRestante"])
        
        for cell in ws_detalle_preg[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFE400", end_color="FFE400", fill_type="solid")

        cx = obtener_conexion()
        try:
            with cx.cursor(pymysql.cursors.DictCursor) as c:
                c.execute("""
                    SELECT rp.id_pregunta, 
                           MIN(rp.pregunta_texto) AS pregunta_texto,
                           rp.opcion_texto,
                           MAX(rp.es_correcta) AS es_correcta,
                           COUNT(*) AS respuestas_recibidas,
                           SUM(rp.es_correcta) AS aciertos,
                           ROUND(AVG(TIME_TO_SEC(rp.tiempo_respuesta)), 2) AS tiempo_prom
                    FROM RESPUESTA_PARTICIPANTE rp
                    WHERE rp.id_partida=%s
                    GROUP BY rp.id_pregunta, rp.opcion_texto
                    ORDER BY rp.id_pregunta, rp.opcion_texto
                """, (id_partida if id_partida else data['header']['id_partida'],))
                rows_preg = c.fetchall()
                
                for r in rows_preg:
                    pct_aciertos = round((int(r['aciertos']) / int(r['respuestas_recibidas']) * 100), 2) if r['respuestas_recibidas'] > 0 else 0
                    ws_detalle_preg.append([
                        r['id_pregunta'],
                        r['pregunta_texto'],
                        r['opcion_texto'],
                        int(r['es_correcta']),
                        int(r['respuestas_recibidas']),
                        int(r['aciertos']),
                        pct_aciertos,
                        float(r['tiempo_prom'] or 0)
                    ])
        finally:
            cx.close()
        _apply_sheet_style(ws_detalle_preg)

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Obtener contenido en bytes
        excel_bytes = output.getvalue()
        
        filename = f"DogeHoot_Reporte_{pin if pin else id_partida}.xlsx"
        
        # Inicializar uploader de Google Drive
        credentials_file = os.path.join(app.root_path, 'credentials.json')
        token_file = os.path.join(app.root_path, 'token_drive.json')
        
        uploader = GoogleDriveUploader(
            credentials_file=credentials_file,
            token_file=token_file
        )
        
        # Buscar o crear carpeta "DogeHoot Reportes"
        carpetas = uploader.listar_archivos(max_resultados=100)
        folder_id = None
        
        if carpetas['success']:
            for archivo in carpetas['files']:
                if archivo['name'] == 'DogeHoot Reportes' and archivo['mimeType'] == 'application/vnd.google-apps.folder':
                    folder_id = archivo['id']
                    break
        
        # Si no existe la carpeta, crearla
        if not folder_id:
            resultado_carpeta = uploader.crear_carpeta('DogeHoot Reportes')
            if resultado_carpeta['success']:
                folder_id = resultado_carpeta['folder_id']
            else:
                return jsonify({
                    "ok": False, 
                    "msg": f"Error al crear carpeta: {resultado_carpeta['message']}"
                }), 500
        
        # Subir archivo desde memoria
        resultado = uploader.subir_desde_memoria(
            contenido=excel_bytes,
            nombre_archivo=filename,
            mime_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            folder_id=folder_id
        )
        
        if not resultado['success']:
            return jsonify({
                "ok": False,
                "msg": f"Error al subir a Drive: {resultado['message']}"
            }), 500
        
        # Compartir el archivo con el correo proporcionado
        file_id = resultado['file_id']
        resultado_compartir = uploader.compartir_archivo(
            file_id=file_id,
            email=email,
            role='reader',  # Solo lectura
            tipo='user'
        )
        
        if resultado_compartir['success']:
            return jsonify({
                "ok": True,
                "msg": f"Reporte subido y compartido exitosamente con {email}",
                "file_name": resultado['file_name'],
                "file_id": resultado['file_id'],
                "web_view_link": resultado['web_view_link'],
                "download_link": resultado.get('download_link'),
                "shared_with": email
            }), 200
        else:
            # El archivo se subió pero no se pudo compartir
            return jsonify({
                "ok": True,
                "msg": f"Reporte subido pero hubo un error al compartir: {resultado_compartir['message']}",
                "file_name": resultado['file_name'],
                "file_id": resultado['file_id'],
                "web_view_link": resultado['web_view_link'],
                "download_link": resultado.get('download_link'),
                "warning": "No se pudo compartir con el correo especificado"
            }), 200
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({"ok": False, "msg": str(e)}), 500


@app.route('/api/report/partida/export-onedrive', methods=['GET'])
def api_report_partida_export_onedrive():
    """
    Genera Excel y lo sube a OneDrive en la carpeta 'DogeHoot/Reportes' y lo comparte con un correo
    
    Uso: 
    - /api/report/partida/export-onedrive?pin=AB12CD&email=correo@ejemplo.com
    - /api/report/partida/export-onedrive?id_partida=123&email=correo@ejemplo.com
    """
    try:
        pin = request.args.get('pin')
        id_partida = request.args.get('id_partida', type=int)
        email = request.args.get('email')
        
        # Validar que se proporcione el correo
        if not email:
            return jsonify({"ok": False, "msg": "Debe proporcionar un correo electrónico (parámetro 'email')"}), 400
        
        # Importar el uploader de OneDrive
        from controladores.onedrive_uploader import OneDriveUploader
        
        # === PASO 1: Obtener datos del reporte (igual que Google Drive) ===
        from controladores import controlador_partidas as c_part
        
        if pin:
            data = c_part.reporte_por_pin(pin)
        elif id_partida:
            data = c_part.reporte_por_partida_id(id_partida)
        else:
            return jsonify({"ok": False, "msg": "Proporcione pin o id_partida"}), 400
        
        if not data:
            return jsonify({"ok": False, "msg": "Partida no encontrada"}), 404
        
        # === PASO 2: Crear el archivo Excel (mismo código que ya tienes) ===
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io
        
        wb = Workbook()
        wb.remove(wb.active)  # Eliminar hoja por defecto

        def _apply_sheet_style(sheet):
            max_col = sheet.max_column or 1
            max_row = sheet.max_row or 1
            center_alignment = Alignment(horizontal="center", vertical="center")
            for col_idx in range(1, max_col + 1):
                sheet.column_dimensions[get_column_letter(col_idx)].width = 25
            for row_idx in range(1, max_row + 1):
                sheet.row_dimensions[row_idx].height = 20
            for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    cell.alignment = center_alignment

        # --- HOJA 1: RESUMEN ---
        ws_resumen = wb.create_sheet("Resumen")

        header_data = data.get("header") or {}
        estado_map = {"F": "Finalizado", "E": "En espera", "P": "En progreso"}
        estado_texto = estado_map.get(header_data.get("estado"), header_data.get("estado") or "")

        def _to_text(value):
            if value is None:
                return ""
            if isinstance(value, datetime):
                return value.strftime("%Y-%m-%d %H:%M:%S")
            return str(value)

        ws_resumen["B1"] = "Datos de la PARTIDA:"
        ws_resumen["B1"].font = Font(bold=True)
        ws_resumen["B1"].alignment = Alignment(horizontal="left", vertical="center")

        info_rows = []
        if header_data:
            info_rows.extend([
                ("ID Partida", _to_text(header_data.get("id_partida"))),
                ("PIN", _to_text(header_data.get("pin"))),
                ("Cuestionario", _to_text(header_data.get("cuestionario_titulo"))),
                ("Estado", estado_texto),
            ])
            inicio = _to_text(header_data.get("fecha_hora_inicio"))
            if inicio:
                info_rows.append(("Fecha de inicio", inicio))
            fin = _to_text(header_data.get("fecha_hora_fin"))
            if fin:
                info_rows.append(("Fecha de finalizacion", fin))

        thin_side = Side(style="thin", color="000000")
        table_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        for row_offset, (label, value) in enumerate(info_rows, start=2):
            label_cell = ws_resumen.cell(row=row_offset, column=2, value=label)
            value_cell = ws_resumen.cell(row=row_offset, column=3, value=value)
            label_cell.font = Font(bold=True)
            label_cell.border = table_border
            value_cell.border = table_border

        header_row_idx = len(info_rows) + 3
        resumen_headers = [
            "Posicion",
            "Usuario",
            "Grupo",
            "PuntajeTotal",
            "Monedas",
            "RespuestasCorrectas",
            "RespuestasIncorrectas",
            "%Acierto",
            "TiempoPromedioResp (s)",
        ]

        for col_idx, value in enumerate(resumen_headers, start=1):
            cell = ws_resumen.cell(row=header_row_idx, column=col_idx, value=value)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFE400", end_color="FFE400", fill_type="solid")

        data_start_row = header_row_idx + 1
        for position, part in enumerate(data["participantes"], start=1):
            row = data_start_row + position - 1
            total_resp = int(part["correctas"]) + int(part["incorrectas"])
            pct_acierto = round((int(part["correctas"]) / total_resp * 100), 2) if total_resp > 0 else 0
            grupo_val = int(part.get("grupo") or 0)
            grupo_texto = grupo_val if grupo_val > 0 else "No"
            ws_resumen.cell(row=row, column=1, value=position)
            ws_resumen.cell(row=row, column=2, value=part["nombre"])
            ws_resumen.cell(row=row, column=3, value=grupo_texto)
            ws_resumen.cell(row=row, column=4, value=int(part["puntaje_total"]))
            ws_resumen.cell(row=row, column=5, value=int(part.get("monedas") or 0))
            ws_resumen.cell(row=row, column=6, value=int(part["correctas"]))
            ws_resumen.cell(row=row, column=7, value=int(part["incorrectas"]))
            ws_resumen.cell(row=row, column=8, value=pct_acierto)
            ws_resumen.cell(row=row, column=9, value=round(float(part["tiempo_prom_seg"] or 0), 2))


        _apply_sheet_style(ws_resumen)
        
        # --- HOJA 2: DETALLE POR PARTICIPANTE ---
        ws_detalle_part = wb.create_sheet("Detalle por participante")
        ws_detalle_part.append(['Usuario', 'Grupo', 'Pregunta N°', 'Correcta (0/1)', 
                               'Tiempo Restante', 'Puntos Otorgados', 'Puntaje Acumulado'])
        
        for cell in ws_detalle_part[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFE400', end_color='FFE400', fill_type='solid')
        
        # Obtener detalle de respuestas
        cx = obtener_conexion()
        try:
            with cx.cursor(pymysql.cursors.DictCursor) as c:
                c.execute("""
                    SELECT pa.nombre, pa.id_grupo, pa.grupo AS numero_grupo, rp.id_pregunta, rp.es_correcta,
                           TIME_TO_SEC(rp.tiempo_respuesta) AS tiempo_seg,
                           rp.puntaje,
                           (SELECT SUM(rp2.puntaje) FROM RESPUESTA_PARTICIPANTE rp2
                            WHERE rp2.id_participante = rp.id_participante 
                            AND rp2.id_respuesta <= rp.id_respuesta) AS puntaje_acum
                    FROM RESPUESTA_PARTICIPANTE rp
                    JOIN PARTICIPANTE pa ON pa.id_participante = rp.id_participante
                    WHERE rp.id_partida = %s
                    ORDER BY pa.nombre, rp.id_pregunta
                """, (id_partida if id_partida else data['header']['id_partida'],))
                rows_part = c.fetchall()
                
                for r in rows_part:
                    ws_detalle_part.append([
                        r['nombre'],
                        (int(r.get('numero_grupo') or 0) or 'No'),
                        r['id_pregunta'],
                        int(r['es_correcta']),
                        int(r['tiempo_seg'] or 0),
                        int(r['puntaje']),
                        int(r['puntaje_acum'] or 0)
                    ])
        finally:
            cx.close()
        _apply_sheet_style(ws_detalle_part)
        
        # --- HOJA 3: DETALLE POR PREGUNTA ---
        ws_detalle_preg = wb.create_sheet("Detalle por pregunta")
        ws_detalle_preg.append(['Pregunta', 'Texto Pregunta', 'Opción', 'Es Correcta (0/1)',
                               'Respuestas Recibidas', 'Aciertos', '% Aciertos', 'Tiempo Promedio'])
        
        for cell in ws_detalle_preg[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFE400', end_color='FFE400', fill_type='solid')
        
        cx = obtener_conexion()
        try:
            with cx.cursor(pymysql.cursors.DictCursor) as c:
                c.execute("""
                    SELECT rp.id_pregunta, MIN(rp.pregunta_texto) AS pregunta_texto,
                           rp.opcion_texto, MAX(rp.es_correcta) AS es_correcta,
                           COUNT(*) AS respuestas_recibidas,
                           SUM(rp.es_correcta) AS aciertos,
                           ROUND(AVG(TIME_TO_SEC(rp.tiempo_respuesta)), 2) AS tiempo_prom
                    FROM RESPUESTA_PARTICIPANTE rp
                    WHERE rp.id_partida = %s
                    GROUP BY rp.id_pregunta, rp.opcion_texto
                    ORDER BY rp.id_pregunta, rp.opcion_texto
                """, (id_partida if id_partida else data['header']['id_partida'],))
                rows_preg = c.fetchall()
                
                for r in rows_preg:
                    pct_aciertos = round(int(r['aciertos']) / int(r['respuestas_recibidas']) * 100, 2) if r['respuestas_recibidas'] > 0 else 0
                    ws_detalle_preg.append([
                        r['id_pregunta'],
                        r['pregunta_texto'],
                        r['opcion_texto'],
                        int(r['es_correcta']),
                        int(r['respuestas_recibidas']),
                        int(r['aciertos']),
                        pct_aciertos,
                        float(r['tiempo_prom'] or 0)
                    ])
        finally:
            cx.close()
        _apply_sheet_style(ws_detalle_preg)
        
        # === PASO 3: Guardar Excel en memoria (bytes) ===
        output = io.BytesIO()
        wb.save(output)
        excel_bytes = output.getvalue()
        
        # Nombre del archivo con timestamp único para evitar conflictos
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'DogeHoot_Reporte_{pin if pin else id_partida}_{timestamp}.xlsx'
        
        # === PASO 4: Subir a OneDrive ===
        uploader = OneDriveUploader()
        
        # Subir el archivo desde memoria a la carpeta DogeHoot/Reportes
        resultado = uploader.subir_desde_memoria(
            excel_bytes,
            filename,
            folder_path='DogeHoot/Reportes'
        )
        
        if not resultado['success']:
            return jsonify({
                'ok': False,
                'msg': f'Error al subir a OneDrive: {resultado["message"]}'
            }), 500
        
        # === PASO 5: Compartir el archivo con el correo proporcionado ===
        file_id = resultado['file_id']
        file_name = resultado['file_name']
        
        resultado_compartir = uploader.compartir_archivo(
            file_id=file_id,
            email=email,
            role='read',  # Solo lectura
            send_notification=True,  # Enviar notificación por correo
            file_name=file_name  # Nombre del archivo para el correo
        )
        
        if resultado_compartir['success']:
            mensaje_exito = f'Reporte subido y enviado exitosamente a {email}'
            if resultado_compartir.get('email_sent'):
                mensaje_exito += ' (correo enviado)'
            else:
                mensaje_exito += ' (enlace creado, revisar correo)'
                
            return jsonify({
                'ok': True,
                'msg': mensaje_exito,
                'file_name': resultado['file_name'],
                'file_id': resultado['file_id'],
                'web_url': resultado['web_url'],  # URL para ver el archivo en OneDrive
                'download_url': resultado.get('download_url'),  # URL de descarga directa
                'share_link': resultado_compartir.get('share_link'),  # Enlace compartido
                'shared_with': email,
                'email_sent': resultado_compartir.get('email_sent', False)
            }), 200
        else:
            # El archivo se subió pero no se pudo compartir
            return jsonify({
                'ok': True,
                'msg': f'Reporte subido pero hubo un error al compartir: {resultado_compartir["message"]}',
                'file_name': resultado['file_name'],
                'file_id': resultado['file_id'],
                'web_url': resultado['web_url'],
                'download_url': resultado.get('download_url'),
                'warning': 'No se pudo compartir con el correo especificado'
            }), 200
            
    except Exception as e:
        traceback.print_exc()
        return jsonify({"ok": False, "msg": str(e)}), 500


# ============================================
# RUTAS PARA ONEDRIVE OAuth
# ============================================

@app.route('/onedrive/auth')
def onedrive_auth():
    """Inicia el proceso de autenticación con OneDrive"""
    uploader = OneDriveUploader()
    auth_url = uploader.get_authorization_url()
    return redirect(auth_url)


@app.route('/onedrive/callback', methods=['GET', 'POST'])
def onedrive_callback():
    """Callback de OneDrive después de la autenticación"""
    # Obtener todos los parámetros para debugging
    all_params = dict(request.args)
    all_form = dict(request.form) if request.form else {}
    
    print("="*50)
    print(f"Método de petición: {request.method}")
    print(f"URL completa: {request.url}")
    print(f"Query params (GET): {all_params}")
    print(f"Form data (POST): {all_form}")
    print(f"Headers: {dict(request.headers)}")
    print("="*50)
    
    # Intentar obtener el código de GET o POST
    code = request.args.get('code') or request.form.get('code')
    error = request.args.get('error') or request.form.get('error')
    error_description = request.args.get('error_description') or request.form.get('error_description')
    
    if error:
        return jsonify({
            'success': False,
            'message': f'Error en autenticación: {error}',
            'error_description': error_description,
            'params_received': all_params,
            'form_received': all_form
        }), 400
    
    if code:
        print(f"✓ Código recibido: {code[:20]}...")
        uploader = OneDriveUploader()
        result = uploader.exchange_code_for_tokens(code)
        
        if result['success']:
            return jsonify({
                'success': True,
                'message': 'OneDrive conectado exitosamente. Ya puedes subir archivos.'
            })
        else:
            return jsonify(result), 400
    
    # Si no hay código ni error, mostrar qué se recibió
    return jsonify({
        'success': False,
        'message': 'No se recibió código de autenticación',
        'params_received': all_params,
        'form_received': all_form,
        'url_received': request.url
    }), 400


@app.route('/onedrive/test-upload')
def onedrive_test_upload():
    """Ruta de prueba para subir un archivo a OneDrive"""
    try:
        uploader = OneDriveUploader()
        
        # Crear un archivo de prueba
        test_content = b"Este es un archivo de prueba de DogeHoot"
        result = uploader.subir_desde_memoria(
            test_content,
            'prueba_dogehoot.txt',
            folder_path='DogeHoot/Pruebas'
        )
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        }), 500

# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# PARTICIPANTES
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------

def _serialize_time_row(row):
    """
    Ayudante para serializar campos de tiempo/fecha en respuestas.
    Usa el mismo patrón que tu _serialize_partida_row.
    """
    if not row:
        return None
    serializable = dict(row)
    
    tiempo = serializable.get("tiempo_respuesta")
    if tiempo and not isinstance(tiempo, str):
        serializable["tiempo_respuesta"] = str(tiempo)

    creado = serializable.get("creado_en")
    if creado and isinstance(creado, datetime):
        serializable["creado_en"] = creado.isoformat()
        
    return serializable

@app.route("/api_registrar_participante", methods=["POST"])
@jwt_required()
def api_registrar_participante():
    """
    [CREATE] Registra un nuevo participante en una partida.
    Inicia el puntaje en 0 por defecto.
    """
    rpta = {"code": 0, "data": {}, "message": ""}
    conexion = None
    try:
        data = request.get_json(silent=True) or {}
        id_partida = data.get("id_partida")
        nombre = data.get("nombre")
        registrado = data.get("registrado", 0) # 0 = invitado, 1 = usuario

        if not id_partida or not nombre:
            rpta["message"] = "id_partida y nombre son obligatorios."
            return jsonify(rpta), 400

        campos = [
            ("id_partida", id_partida),
            ("nombre", nombre),
            ("registrado", registrado),
            ("puntaje", data.get("puntaje", 0)),
            ("id_usuario", data.get("id_usuario")), # Opcional
            ("id_grupo", data.get("id_grupo")),     # Opcional
            ("grupo", data.get("grupo")),         # Opcional
        ]

        columnas = ", ".join(col for col, val in campos if val is not None)
        valores = [val for _, val in campos if val is not None]
        placeholders = ", ".join(["%s"] * len(valores))

        conexion = obtener_conexion()
        with conexion.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute(
                f"INSERT INTO PARTICIPANTE ({columnas}) VALUES ({placeholders})",
                valores
            )
            nuevo_id = cursor.lastrowid
            conexion.commit()
            
            cursor.execute("SELECT * FROM PARTICIPANTE WHERE id_participante = %s", (nuevo_id,))
            nuevo_participante = cursor.fetchone()

        rpta["code"] = 1
        rpta["data"] = nuevo_participante
        rpta["message"] = "Participante creado correctamente"
        return jsonify(rpta), 201
        
    except pymysql.Error as e:
        if conexion:
            conexion.rollback()
        rpta["message"] = f"No se pudo crear el participante: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()

@app.route("/api_obtener_participantes", methods=["GET"])
@jwt_required()
def api_obtener_participantes():
    """
    [READ ALL] Obtiene todos los participantes.
    Filtro opcional: /api_obtener_participantes?id_partida=123
    """
    rpta = {"code": 0, "data": {}, "message": ""}
    conexion = None
    try:
        id_partida_filtro = request.args.get("id_partida", type=int)

        conexion = obtener_conexion()
        with conexion.cursor(pymysql.cursors.DictCursor) as cursor:
            if id_partida_filtro:
                cursor.execute(
                    "SELECT * FROM PARTICIPANTE WHERE id_partida = %s ORDER BY puntaje DESC",
                    (id_partida_filtro,)
                )
            else:
                cursor.execute("SELECT * FROM PARTICIPANTE ORDER BY id_partida DESC, puntaje DESC")
            
            participantes = cursor.fetchall()

        rpta["code"] = 1
        rpta["data"] = participantes
        rpta["message"] = "Participantes obtenidos correctamente"
        return jsonify(rpta)
        
    except Exception as e:
        rpta["message"] = f"No se pudieron obtener los participantes: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()

@app.route("/api_obtener_participante_por_id/<int:id_participante>", methods=["GET"])
@jwt_required()
def api_obtener_participante_por_id(id_participante):
    """
    [READ BY ID] Obtiene un participante específico por su ID.
    """
    rpta = {"code": 0, "data": {}, "message": ""}
    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("SELECT * FROM PARTICIPANTE WHERE id_participante = %s", (id_participante,))
            participante = cursor.fetchone()

        if not participante:
            rpta["message"] = "Participante no encontrado"
            return jsonify(rpta), 404

        rpta["code"] = 1
        rpta["data"] = participante
        rpta["message"] = "Participante obtenido correctamente"
        return jsonify(rpta)
        
    except Exception as e:
        rpta["message"] = f"No se pudo obtener el participante: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()

@app.route("/api_actualizar_participante/<int:id_participante>", methods=["PUT"])
@jwt_required()
def api_actualizar_participante(id_participante):
    """
    [UPDATE] Actualiza los campos de un participante.
    """
    rpta = {"code": 0, "data": {}, "message": ""}
    data = request.get_json(silent=True) or {}
    
    campos_validos = {"nombre", "registrado", "puntaje", "id_usuario", "id_grupo", "grupo"}
    sets = []
    valores = []
    
    for campo in campos_validos:
        if campo in data:
            sets.append(f"{campo} = %s")
            valores.append(data[campo])

    if not sets:
        rpta["message"] = "No se enviaron campos a actualizar."
        return jsonify(rpta), 400

    valores.append(id_participante)
    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute(
                f"UPDATE PARTICIPANTE SET {', '.join(sets)} WHERE id_participante = %s",
                valores
            )
            
            if cursor.rowcount == 0:
                rpta["message"] = "Datos a actualizar son iguales o participante no encontrado."
                return jsonify(rpta), 404
            
            conexion.commit()
            
            cursor.execute("SELECT * FROM PARTICIPANTE WHERE id_participante = %s", (id_participante,))
            participante_actualizado = cursor.fetchone()

        rpta["code"] = 1
        rpta["data"] = participante_actualizado
        rpta["message"] = "Participante actualizado correctamente"
        return jsonify(rpta)
        
    except Exception as e:
        if conexion:
            conexion.rollback()
        rpta["message"] = f"No se pudo actualizar el participante: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()

@app.route("/api_eliminar_participante/<int:id_participante>", methods=["DELETE"])
@jwt_required()
def api_eliminar_participante(id_participante):
    """
    [DELETE] Elimina un participante Y TODAS SUS RESPUESTAS ASOCIADAS.
    Usa una transacción para mantener la integridad.
    """
    rpta = {"code": 0, "data": {}, "message": ""}
    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor() as cursor:
            cursor.execute("SELECT 1 FROM PARTICIPANTE WHERE id_participante = %s", (id_participante,))
            if not cursor.fetchone():
                 rpta["message"] = "Participante no encontrado"
                 return jsonify(rpta), 404

            cursor.execute("DELETE FROM RESPUESTA_PARTICIPANTE WHERE id_participante = %s", (id_participante,))
            respuestas_borradas = cursor.rowcount

            cursor.execute("DELETE FROM PARTICIPANTE WHERE id_participante = %s", (id_participante,))
            participante_borrado = cursor.rowcount
            
            conexion.commit()

        rpta["code"] = 1
        rpta["data"] = {"respuestas_eliminadas": respuestas_borradas, "participantes_eliminados": participante_borrado}
        rpta["message"] = "Participante y sus respuestas eliminados correctamente"
        return jsonify(rpta)
        
    except pymysql.Error as e:
        if conexion:
            conexion.rollback()
        # Error de FK si está siendo referenciado por GRUPO
        if e.errno == 1451: 
             rpta["message"] = f"No se pudo eliminar: El participante es líder de un grupo (id_grupo). Primero elimine el grupo."
             return jsonify(rpta), 409
        rpta["message"] = f"No se pudo eliminar el participante: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()


# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# RESPUESTA_PARTICIPANTE
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------

@app.route("/api_registrar_respuesta", methods=["POST"])
@jwt_required()
def api_registrar_respuesta():
    """
    [CREATE] Registra una nueva respuesta Y ACTUALIZA EL PUNTAJE del participante.
    Esta es una API de admin, asume que los datos son correctos.
    La lógica del juego debe usar la ruta '/api/game/answer'
    """
    rpta = {"code": 0, "data": {}, "message": ""}
    conexion = None
    try:
        data = request.get_json(silent=True) or {}
        
        id_participante = data.get("id_participante")
        id_partida = data.get("id_partida")
        id_pregunta = data.get("id_pregunta")
        puntaje = data.get("puntaje", 0)
        
        if not all([id_participante, id_partida, id_pregunta]):
            rpta["message"] = "id_participante, id_partida, y id_pregunta son obligatorios."
            return jsonify(rpta), 400
        
        # Campos opcionales/calculados
        campos = [
            ("id_participante", id_participante),
            ("id_opcion", data.get("id_opcion")),
            ("id_partida", id_partida),
            ("id_pregunta", id_pregunta),
            ("pregunta_texto", data.get("pregunta_texto", "")),
            ("opcion_texto", data.get("opcion_texto", "")),
            ("es_correcta", data.get("es_correcta", 0)),
            ("puntaje", puntaje),
            ("tiempo_respuesta", data.get("tiempo_respuesta", "00:00:00")),
        ]

        columnas = ", ".join(col for col, _ in campos)
        valores = [val for _, val in campos]
        placeholders = ", ".join(["%s"] * len(valores))

        conexion = obtener_conexion()
        with conexion.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute(
                f"INSERT INTO RESPUESTA_PARTICIPANTE ({columnas}) VALUES ({placeholders})",
                valores
            )
            nuevo_id = cursor.lastrowid

            cursor.execute(
                "UPDATE PARTICIPANTE SET puntaje = puntaje + %s WHERE id_participante = %s",
                (puntaje, id_participante)
            )
            
            conexion.commit()
            
            cursor.execute("SELECT * FROM RESPUESTA_PARTICIPANTE WHERE id_respuesta = %s", (nuevo_id,))
            nueva_respuesta = _serialize_time_row(cursor.fetchone())

        rpta["code"] = 1
        rpta["data"] = nueva_respuesta
        rpta["message"] = "Respuesta registrada y puntaje actualizado"
        return jsonify(rpta), 201
        
    except pymysql.Error as e:
        if conexion:
            conexion.rollback()
        error_code = e.args[0] 
        
        if error_code == 1062: # Error de 'uniq_resp' (Llave duplicada)
            rpta["message"] = f"Error de duplicado: El participante {id_participante} ya respondió la pregunta {id_pregunta}."
            return jsonify(rpta), 409
        
        if error_code == 1452: # Error de 'Foreign Key'
            rpta["message"] = f"Error de integridad (FK): El 'id_participante' ({id_participante}) o el 'id_opcion' no existen en la base de datos."
            return jsonify(rpta), 400
        rpta["message"] = f"No se pudo crear la respuesta: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()

@app.route("/api_obtener_respuestas", methods=["GET"])
@jwt_required()
def api_obtener_respuestas():
    """
    [READ ALL] Obtiene todas las respuestas.
    Filtros opcionales:
    /api_obtener_respuestas?id_partida=123
    /api_obtener_respuestas?id_participante=456
    """
    rpta = {"code": 0, "data": {}, "message": ""}
    conexion = None
    try:
        id_partida_filtro = request.args.get("id_partida", type=int)
        id_participante_filtro = request.args.get("id_participante", type=int)

        query = "SELECT * FROM RESPUESTA_PARTICIPANTE"
        params = []
        
        if id_partida_filtro:
            query += " WHERE id_partida = %s"
            params.append(id_partida_filtro)
        elif id_participante_filtro:
            query += " WHERE id_participante = %s"
            params.append(id_participante_filtro)
            
        query += " ORDER BY creado_en DESC"

        conexion = obtener_conexion()
        with conexion.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute(query, params)
            respuestas = [_serialize_time_row(row) for row in cursor.fetchall()]

        rpta["code"] = 1
        rpta["data"] = respuestas
        rpta["message"] = "Respuestas obtenidas correctamente"
        return jsonify(rpta)
        
    except Exception as e:
        rpta["message"] = f"No se pudieron obtener las respuestas: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()

@app.route("/api_obtener_respuesta_por_id/<int:id_respuesta>", methods=["GET"])
@jwt_required()
def api_obtener_respuesta_por_id(id_respuesta):
    """
    [READ BY ID] Obtiene una respuesta específica por su ID.
    """
    rpta = {"code": 0, "data": {}, "message": ""}
    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute("SELECT * FROM RESPUESTA_PARTICIPANTE WHERE id_respuesta = %s", (id_respuesta,))
            respuesta = _serialize_time_row(cursor.fetchone())

        if not respuesta:
            rpta["message"] = "Respuesta no encontrada"
            return jsonify(rpta), 404

        rpta["code"] = 1
        rpta["data"] = respuesta
        rpta["message"] = "Respuesta obtenida correctamente"
        return jsonify(rpta)
        
    except Exception as e:
        rpta["message"] = f"No se pudo obtener la respuesta: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()

@app.route("/api_actualizar_respuesta/<int:id_respuesta>", methods=["PUT"])
@jwt_required()
def api_actualizar_respuesta(id_respuesta):
    """
    [UPDATE] Actualiza una respuesta.
    ¡CUIDADO! Esto recalcula el puntaje del participante.
    """
    rpta = {"code": 0, "data": {}, "message": ""}
    data = request.get_json(silent=True) or {}
    
    campos_validos = {
        "id_opcion", "pregunta_texto", "opcion_texto", 
        "es_correcta", "puntaje", "tiempo_respuesta"
    }
    sets = []
    valores = []
    
    for campo in campos_validos:
        if campo in data:
            sets.append(f"{campo} = %s")
            valores.append(data[campo])

    if not sets:
        rpta["message"] = "No se enviaron campos a actualizar."
        return jsonify(rpta), 400

    valores.append(id_respuesta)
    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute(
                "SELECT id_participante, puntaje FROM RESPUESTA_PARTICIPANTE WHERE id_respuesta = %s",
                (id_respuesta,)
            )
            respuesta_antigua = cursor.fetchone()
            if not respuesta_antigua:
                rpta["message"] = "Respuesta no encontrada"
                return jsonify(rpta), 404
            
            puntaje_antiguo = respuesta_antigua.get("puntaje", 0)
            id_participante = respuesta_antigua.get("id_participante")
            
            cursor.execute(
                "UPDATE PARTICIPANTE SET puntaje = puntaje - %s WHERE id_participante = %s",
                (puntaje_antiguo, id_participante)
            )
            
            cursor.execute(
                f"UPDATE RESPUESTA_PARTICIPANTE SET {', '.join(sets)} WHERE id_respuesta = %s",
                valores
            )
            
            puntaje_nuevo = data.get("puntaje", puntaje_antiguo)
            cursor.execute(
                "UPDATE PARTICIPANTE SET puntaje = puntaje + %s WHERE id_participante = %s",
                (puntaje_nuevo, id_participante)
            )

            conexion.commit()

            cursor.execute("SELECT * FROM RESPUESTA_PARTICIPANTE WHERE id_respuesta = %s", (id_respuesta,))
            respuesta_actualizada = _serialize_time_row(cursor.fetchone())

        rpta["code"] = 1
        rpta["data"] = respuesta_actualizada
        rpta["message"] = "Respuesta actualizada y puntaje recalculado"
        return jsonify(rpta)
        
    except Exception as e:
        if conexion:
            conexion.rollback()
        rpta["message"] = f"No se pudo actualizar la respuesta: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()

@app.route("/api_eliminar_respuesta/<int:id_respuesta>", methods=["DELETE"])
@jwt_required()
def api_eliminar_respuesta(id_respuesta):
    """
    [DELETE] Elimina una respuesta Y RESTA EL PUNTAJE del participante.
    Usa una transacción para mantener la integridad.
    """
    rpta = {"code": 0, "data": {}, "message": ""}
    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor(pymysql.cursors.DictCursor) as cursor:
            cursor.execute(
                "SELECT id_participante, puntaje FROM RESPUESTA_PARTICIPANTE WHERE id_respuesta = %s",
                (id_respuesta,)
            )
            respuesta = cursor.fetchone()
            if not respuesta:
                 rpta["message"] = "Respuesta no encontrada"
                 return jsonify(rpta), 404

            puntaje_a_restar = respuesta.get("puntaje", 0)
            id_participante = respuesta.get("id_participante")

            cursor.execute("DELETE FROM RESPUESTA_PARTICIPANTE WHERE id_respuesta = %s", (id_respuesta,))

            if id_participante and puntaje_a_restar > 0:
                cursor.execute(
                    "UPDATE PARTICIPANTE SET puntaje = puntaje - %s WHERE id_participante = %s",
                    (puntaje_a_restar, id_participante)
                )

            conexion.commit()

        rpta["code"] = 1
        rpta["data"] = {"puntaje_restado": puntaje_a_restar, "id_participante_afectado": id_participante}
        rpta["message"] = "Respuesta eliminada y puntaje de participante actualizado"
        return jsonify(rpta)
        
    except pymysql.Error as e:
        if conexion:
            conexion.rollback()
        rpta["message"] = f"No se pudo eliminar la respuesta: {e}"
        return jsonify(rpta), 500
    finally:
        if conexion:
            conexion.close()
            
            

# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# ENCRIPTACIÓN
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------
@app.route('/api_test_encriptar_todo')
def api_test_encriptar_todo():
    conexion = obtener_conexion()
    try:
        with conexion.cursor() as cursor:  # ✅ Solo un cursor
            # 1. Obtener usuarios
            cursor.execute("""
                SELECT id_usuario, correo, nombre_usuario, contraseña 
                FROM USUARIO 
                WHERE VIGENTE = true AND LENGTH(contraseña) < 50
            """)
            usuarios = cursor.fetchall()
            
            if not usuarios:
                return jsonify({
                    "Status": 0,
                    "Mensaje": "No se encontraron usuarios para encriptar sus contraseñas"
                })
            
            # 2. Actualizar contraseñas
            usuarios_actualizados = 0
            for usuario in usuarios:
                id_usuario = usuario['id_usuario']
                passenc = encriptar_sha256(usuario['contraseña'])
                
                cursor.execute(
                    "UPDATE USUARIO SET contraseña = %s WHERE id_usuario = %s",
                    (passenc, id_usuario)
                )
                usuarios_actualizados += 1
            
            # 3. Commit una sola vez al final
            conexion.commit()
            
            # ✅ Return FUERA del for
            return jsonify({
                "Status": 1,
                "Mensaje": f"{usuarios_actualizados} contraseñas encriptadas correctamente"
            })
            
    except Exception as e:
        if conexion:
            conexion.rollback()
        return jsonify({
            "Status": 0,
            "Mensaje": f"Error al encriptar contraseñas: {str(e)}"
        }), 500
    finally:
        if conexion:
            conexion.close()
    
