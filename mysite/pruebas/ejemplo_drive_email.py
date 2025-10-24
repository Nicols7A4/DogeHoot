"""
EJEMPLO COMPLETO - Subir a Drive y Enviar Email
================================================
Este ejemplo muestra cómo:
1. Crear un reporte Excel
2. Subirlo a Google Drive
3. Enviar el link por correo electrónico

Autor: DogeHoot Team
Fecha: 23 de octubre de 2025
"""

import sys
import os
from datetime import datetime

# Agregar el directorio padre al path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from controladores.google_drive_uploader import GoogleDriveUploader
from controladores.email_sender import EmailSender


def crear_reporte_ejemplo():
    """Crea un reporte Excel de ejemplo"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Partidas"
        
        # Estilos
        header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Encabezados
        headers = ["ID Partida", "Anfitrión", "Fecha", "Jugadores", "Duración", "Puntuación Máxima"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Datos de ejemplo
        datos = [
            [1, "Juan Pérez", "2025-10-20", 15, "25 min", 1500],
            [2, "María García", "2025-10-21", 12, "20 min", 1350],
            [3, "Carlos López", "2025-10-22", 18, "30 min", 1750],
            [4, "Ana Martínez", "2025-10-23", 10, "18 min", 1200],
        ]
        
        for row_idx, row_data in enumerate(datos, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 18
        
        # Guardar
        archivo = f"reporte_dogehoot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(archivo)
        print(f"✅ Reporte creado: {archivo}")
        return archivo
        
    except Exception as e:
        print(f"❌ Error al crear reporte: {e}")
        return None


def ejemplo_completo_drive_email():
    """Ejemplo completo: Crear reporte, subir a Drive y enviar email"""
    
    print("\n" + "="*70)
    print("🎮 DOGEHOOT - EJEMPLO COMPLETO: DRIVE + EMAIL")
    print("="*70 + "\n")
    
    # Paso 1: Crear reporte
    print("📊 Paso 1: Creando reporte Excel...")
    archivo_excel = crear_reporte_ejemplo()
    
    if not archivo_excel:
        print("❌ No se pudo crear el reporte")
        return
    
    # Paso 2: Subir a Google Drive
    print("\n📤 Paso 2: Subiendo a Google Drive...")
    try:
        uploader = GoogleDriveUploader(credentials_file='../credentials.json')
        
        resultado_drive = uploader.subir_archivo(
            ruta_archivo=archivo_excel,
            nombre_drive=f"DogeHoot_Reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not resultado_drive['success']:
            print(f"❌ Error al subir a Drive: {resultado_drive['message']}")
            return
        
        print(f"✅ Archivo subido exitosamente")
        print(f"   🆔 ID: {resultado_drive['file_id']}")
        print(f"   🔗 Link: {resultado_drive['web_view_link']}")
        
        file_id = resultado_drive['file_id']
        web_link = resultado_drive['web_view_link']
        
    except Exception as e:
        print(f"❌ Error: {e}")
        return
    
    # Paso 3: Compartir archivo públicamente
    print("\n🔓 Paso 3: Compartiendo archivo públicamente...")
    try:
        resultado_compartir = uploader.compartir_archivo(
            file_id=file_id,
            tipo='anyone',
            role='reader'
        )
        
        if resultado_compartir['success']:
            print("✅ Archivo compartido públicamente")
        else:
            print(f"⚠️ Advertencia: {resultado_compartir['message']}")
    
    except Exception as e:
        print(f"⚠️ No se pudo compartir: {e}")
    
    # Paso 4: Enviar email con el link
    print("\n📧 Paso 4: Enviando correo electrónico...")
    
    # Solicitar email del destinatario
    destinatario = input("   👤 Email del destinatario (Enter para omitir): ").strip()
    
    if not destinatario:
        print("   ⏭️ Envío de email omitido")
    else:
        try:
            email_sender = EmailSender()
            
            # Crear mensaje HTML
            mensaje_html = f"""
            <html>
                <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                    <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
                        <h2 style="color: #4A90E2; text-align: center;">🎮 DogeHoot - Tu Reporte está Listo</h2>
                        <hr style="border: 1px solid #eee;">
                        
                        <p>¡Hola!</p>
                        
                        <p>Tu reporte de partidas de DogeHoot ha sido generado y está disponible en Google Drive.</p>
                        
                        <div style="background-color: #f5f5f5; padding: 20px; border-radius: 5px; margin: 20px 0; text-align: center;">
                            <p style="margin-bottom: 15px;"><strong>📊 Reporte de Partidas</strong></p>
                            <a href="{web_link}" 
                               style="display: inline-block; padding: 12px 30px; background-color: #4A90E2; color: white; text-decoration: none; border-radius: 5px; font-weight: bold;">
                                📥 Abrir en Google Drive
                            </a>
                        </div>
                        
                        <div style="background-color: #fff3cd; padding: 15px; border-radius: 5px; margin: 20px 0; border-left: 4px solid #ffc107;">
                            <p style="margin: 0;"><strong>💡 Nota:</strong> El archivo es de solo lectura. Si deseas editarlo, haz clic en "Hacer una copia" en Google Drive.</p>
                        </div>
                        
                        <p style="margin-top: 20px;">Información del reporte:</p>
                        <ul>
                            <li><strong>📅 Fecha de generación:</strong> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</li>
                            <li><strong>📁 Archivo:</strong> {resultado_drive['file_name']}</li>
                            <li><strong>🔗 Link directo:</strong> <a href="{web_link}">{web_link[:50]}...</a></li>
                        </ul>
                        
                        <p>Gracias por usar DogeHoot 🐕</p>
                        
                        <hr style="border: 1px solid #eee;">
                        <p style="font-size: 12px; color: #999; text-align: center;">
                            Este correo fue generado automáticamente. Por favor no respondas a este mensaje.
                        </p>
                    </div>
                </body>
            </html>
            """
            
            resultado_email = email_sender.enviar_correo_gmail(
                destinatario=destinatario,
                asunto='📊 DogeHoot - Tu Reporte en Google Drive',
                mensaje_html=mensaje_html
            )
            
            if resultado_email['success']:
                print(f"   ✅ Email enviado exitosamente a {destinatario}")
            else:
                print(f"   ❌ Error al enviar email: {resultado_email['message']}")
        
        except Exception as e:
            print(f"   ❌ Error: {e}")
    
    # Paso 5: Limpiar archivo local
    print("\n🗑️  Paso 5: Limpiando archivos locales...")
    try:
        os.remove(archivo_excel)
        print(f"   ✅ Archivo local eliminado: {archivo_excel}")
    except Exception as e:
        print(f"   ⚠️ No se pudo eliminar: {e}")
    
    # Resumen final
    print("\n" + "="*70)
    print("✅ PROCESO COMPLETADO")
    print("="*70)
    print(f"\n📋 Resumen:")
    print(f"   📊 Reporte creado y subido a Google Drive")
    print(f"   🔗 Link: {web_link}")
    print(f"   📧 Email enviado: {'Sí' if destinatario else 'No'}")
    print("\n" + "="*70 + "\n")


def ejemplo_solo_drive():
    """Ejemplo simple: Solo subir un archivo a Drive"""
    
    print("\n" + "="*70)
    print("📤 EJEMPLO SIMPLE: SUBIR ARCHIVO A DRIVE")
    print("="*70 + "\n")
    
    # Crear archivo de prueba
    archivo = "test_simple.txt"
    with open(archivo, 'w', encoding='utf-8') as f:
        f.write(f"Prueba de DogeHoot - {datetime.now()}\n")
    
    print(f"📄 Archivo creado: {archivo}")
    
    # Subir a Drive
    try:
        uploader = GoogleDriveUploader(credentials_file='../credentials.json')
        resultado = uploader.subir_archivo(archivo)
        
        if resultado['success']:
            print(f"\n✅ ¡ÉXITO!")
            print(f"   📋 Nombre: {resultado['file_name']}")
            print(f"   🔗 Ver en Drive: {resultado['web_view_link']}")
        else:
            print(f"\n❌ Error: {resultado['message']}")
        
        # Limpiar
        os.remove(archivo)
        
    except Exception as e:
        print(f"\n❌ Error: {e}")
    
    print("\n" + "="*70 + "\n")


if __name__ == "__main__":
    print("\n🎮 DOGEHOOT - EJEMPLOS DE USO")
    print("\nSelecciona un ejemplo:")
    print("1. Ejemplo completo (Reporte + Drive + Email)")
    print("2. Ejemplo simple (Solo subir archivo)")
    print("0. Salir")
    
    opcion = input("\n👉 Opción: ").strip()
    
    if opcion == '1':
        ejemplo_completo_drive_email()
    elif opcion == '2':
        ejemplo_solo_drive()
    elif opcion == '0':
        print("\n👋 ¡Hasta luego!\n")
    else:
        print("\n❌ Opción inválida\n")
