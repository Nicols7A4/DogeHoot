"""
SCRIPT DE PRUEBA - Google Drive Uploader
=========================================
Prueba la subida de archivos a Google Drive
Compatible con PC Local y PythonAnywhere

INSTRUCCIONES:
1. PRIMERA VEZ EN PC LOCAL:
   - Ejecuta este script
   - Se abrirá navegador para autenticarte
   - Se generará token_drive.json
   
2. PARA PYTHONANYWHERE:
   - Sube credentials.json y token_drive.json a /home/usuario/DogeHoot-2/mysite/
   - Ejecuta este mismo script (detecta automáticamente el entorno)

Autor: DogeHoot Team
Fecha: 23 de octubre de 2025
"""

import sys
import os

# ==================== DETECCIÓN DE ENTORNO ====================

def es_pythonanywhere():
    """Detecta si estamos ejecutando en PythonAnywhere"""
    # PythonAnywhere tiene estas variables de entorno
    return (
        'PYTHONANYWHERE_SITE' in os.environ or 
        'PYTHONANYWHERE_DOMAIN' in os.environ or
        '/home/' in os.path.abspath(__file__)
    )

def obtener_configuracion():
    """Obtiene la configuración según el entorno"""
    if es_pythonanywhere():
        # Configuración para PythonAnywhere
        user = os.environ.get('USER', 'tu_usuario')
        base_dir = f'/home/{user}/DogeHoot-2/mysite'
        
        config = {
            'entorno': 'PythonAnywhere',
            'icono': '🌐',
            'base_dir': base_dir,
            'credentials_file': os.path.join(base_dir, 'credentials.json'),
            'token_file': os.path.join(base_dir, 'token_drive.json'),
            'usar_rutas_absolutas': True
        }
    else:
        # Configuración para PC Local
        base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
        
        config = {
            'entorno': 'PC Local',
            'icono': '💻',
            'base_dir': base_dir,
            'credentials_file': os.path.join(base_dir, 'credentials.json'),
            'token_file': os.path.join(base_dir, 'token_drive.json'),
            'usar_rutas_absolutas': False
        }
    
    return config

# Obtener configuración del entorno
CONFIG = obtener_configuracion()

# Agregar el directorio base al path
sys.path.insert(0, CONFIG['base_dir'])

from controladores.google_drive_uploader import GoogleDriveUploader, subir_archivo_rapido


# ==================== VERIFICACIÓN DE ARCHIVOS ====================

def verificar_archivos_necesarios():
    """Verifica que existan credentials.json y token_drive.json"""
    print(f"\n{CONFIG['icono']} Entorno detectado: {CONFIG['entorno']}")
    print(f"📂 Directorio base: {CONFIG['base_dir']}\n")
    
    archivos_faltantes = []
    
    # Verificar credentials.json
    if os.path.exists(CONFIG['credentials_file']):
        print(f"✅ credentials.json encontrado")
    else:
        print(f"❌ credentials.json NO encontrado en: {CONFIG['credentials_file']}")
        archivos_faltantes.append('credentials.json')
    
    # Verificar token_drive.json
    if os.path.exists(CONFIG['token_file']):
        print(f"✅ token_drive.json encontrado")
    else:
        print(f"⚠️  token_drive.json NO encontrado en: {CONFIG['token_file']}")
        if es_pythonanywhere():
            print(f"   → Debes generarlo en tu PC local primero y subirlo aquí")
            archivos_faltantes.append('token_drive.json')
        else:
            print(f"   → Se generará automáticamente al autenticarte")
    
    if archivos_faltantes and es_pythonanywhere():
        print(f"\n❌ Archivos faltantes: {', '.join(archivos_faltantes)}")
        print(f"\n📋 Pasos para resolver:")
        print(f"   1. Ejecuta este script en tu PC local primero")
        print(f"   2. Sube credentials.json y token_drive.json a PythonAnywhere")
        print(f"   3. Colócalos en: {CONFIG['base_dir']}/")
        return False
    
    print()
    return True


def probar_subida_basica():
    """Prueba básica de subida de archivo"""
    print("\n" + "="*60)
    print("🚀 PRUEBA 1: Subida Básica de Archivo")
    print("="*60)
    
    # Crear un archivo de prueba en ubicación apropiada
    if CONFIG['usar_rutas_absolutas']:
        archivo_prueba = os.path.join(CONFIG['base_dir'], 'pruebas', 'test_drive.txt')
    else:
        archivo_prueba = "test_drive.txt"
    
    with open(archivo_prueba, 'w', encoding='utf-8') as f:
        f.write(f"¡Hola desde DogeHoot! 🐕\n")
        f.write(f"Este es un archivo de prueba para Google Drive\n")
        f.write(f"Entorno: {CONFIG['entorno']}\n")
        f.write(f"Fecha: 23 de octubre de 2025\n")
    
    print(f"📄 Archivo de prueba creado: {archivo_prueba}")
    
    # Subir a Drive
    try:
        resultado = subir_archivo_rapido(
            ruta_archivo=archivo_prueba,
            nombre_drive="DogeHoot_Test.txt",
            credentials_file=CONFIG['credentials_file']
        )
        
        if resultado['success']:
            print("✅ ÉXITO!")
            print(f"   📋 Archivo: {resultado['file_name']}")
            print(f"   🆔 ID: {resultado['file_id']}")
            print(f"   🔗 Ver en Drive: {resultado['web_view_link']}")
            
            # Limpiar archivo de prueba
            os.remove(archivo_prueba)
            print(f"\n🗑️  Archivo local eliminado: {archivo_prueba}")
        else:
            print(f"❌ ERROR: {resultado['message']}")
            
    except Exception as e:
        print(f"❌ EXCEPCIÓN: {str(e)}")
        import traceback
        traceback.print_exc()
    
    print("="*60 + "\n")


def probar_subida_excel():
    """Prueba subida de archivo Excel"""
    print("\n" + "="*60)
    print("📊 PRUEBA 2: Subida de Archivo Excel")
    print("="*60)
    
    try:
        from openpyxl import Workbook
        
        # Crear Excel de prueba
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte DogeHoot"
        
        # Agregar datos
        ws['A1'] = "ID"
        ws['B1'] = "Jugador"
        ws['C1'] = "Puntuación"
        ws['A2'] = 1
        ws['B2'] = "Jugador 1"
        ws['C2'] = 100
        ws['A3'] = 2
        ws['B3'] = "Jugador 2"
        ws['C3'] = 85
        
        # Guardar en ubicación apropiada
        if CONFIG['usar_rutas_absolutas']:
            archivo_excel = os.path.join(CONFIG['base_dir'], 'pruebas', 'reporte_prueba_dogehoot.xlsx')
        else:
            archivo_excel = "reporte_prueba_dogehoot.xlsx"
        
        wb.save(archivo_excel)
        print(f"📊 Archivo Excel creado: {archivo_excel}")
        
        # Subir a Drive
        uploader = GoogleDriveUploader(
            credentials_file=CONFIG['credentials_file'],
            token_file=CONFIG['token_file']
        )
        
        resultado = uploader.subir_archivo(
            ruta_archivo=archivo_excel,
            nombre_drive="DogeHoot_Reporte_Prueba.xlsx"
        )
        
        if resultado['success']:
            print("✅ ÉXITO!")
            print(f"   📋 Archivo: {resultado['file_name']}")
            print(f"   🆔 ID: {resultado['file_id']}")
            print(f"   🔗 Ver en Drive: {resultado['web_view_link']}")
            
            # Limpiar
            os.remove(archivo_excel)
            print(f"\n🗑️  Archivo local eliminado: {archivo_excel}")
        else:
            print(f"❌ ERROR: {resultado['message']}")
            
    except ImportError:
        print("❌ ERROR: Instala openpyxl con: pip install openpyxl")
    except Exception as e:
        print(f"❌ EXCEPCIÓN: {str(e)}")
        import traceback
        traceback.print_exc()
    
    print("="*60 + "\n")


def probar_crear_carpeta():
    """Prueba creación de carpeta"""
    print("\n" + "="*60)
    print("📁 PRUEBA 3: Crear Carpeta en Drive")
    print("="*60)
    
    try:
        uploader = GoogleDriveUploader(
            credentials_file=CONFIG['credentials_file'],
            token_file=CONFIG['token_file']
        )
        resultado = uploader.crear_carpeta("DogeHoot Reportes")
        
        if resultado['success']:
            print("✅ ÉXITO!")
            print(f"   📁 Carpeta: {resultado['folder_name']}")
            print(f"   🆔 ID: {resultado['folder_id']}")
            print(f"   🔗 Ver en Drive: {resultado['web_view_link']}")
            return resultado['folder_id']
        else:
            print(f"❌ ERROR: {resultado['message']}")
            return None
            
    except Exception as e:
        print(f"❌ EXCEPCIÓN: {str(e)}")
        import traceback
        traceback.print_exc()
        return None
    
    print("="*60 + "\n")


def probar_listar_archivos():
    """Prueba listar archivos"""
    print("\n" + "="*60)
    print("📋 PRUEBA 4: Listar Archivos en Drive")
    print("="*60)
    
    try:
        uploader = GoogleDriveUploader(
            credentials_file=CONFIG['credentials_file'],
            token_file=CONFIG['token_file']
        )
        resultado = uploader.listar_archivos(max_resultados=5)
        
        if resultado['success']:
            print(f"✅ {resultado['message']}\n")
            
            if resultado['files']:
                for i, archivo in enumerate(resultado['files'], 1):
                    print(f"{i}. 📄 {archivo['name']}")
                    print(f"   🆔 ID: {archivo['id']}")
                    print(f"   📅 Creado: {archivo.get('createdTime', 'N/A')}")
                    print(f"   🔗 Link: {archivo.get('webViewLink', 'N/A')}")
                    print()
            else:
                print("   (No hay archivos)")
        else:
            print(f"❌ ERROR: {resultado['message']}")
            
    except Exception as e:
        print(f"❌ EXCEPCIÓN: {str(e)}")
        import traceback
        traceback.print_exc()
    
    print("="*60 + "\n")


def menu_principal():
    """Menú principal de pruebas"""
    print("\n" + "="*60)
    print("🐕 DOGEHOOT - PRUEBA DE GOOGLE DRIVE UPLOADER")
    print("="*60)
    print("\nSelecciona una opción:")
    print("1. Probar subida básica de archivo")
    print("2. Probar subida de archivo Excel")
    print("3. Crear carpeta en Drive")
    print("4. Listar archivos en Drive")
    print("5. Ejecutar todas las pruebas")
    print("0. Salir")
    print("="*60)
    
    opcion = input("\n👉 Opción: ").strip()
    
    if opcion == '1':
        probar_subida_basica()
    elif opcion == '2':
        probar_subida_excel()
    elif opcion == '3':
        probar_crear_carpeta()
    elif opcion == '4':
        probar_listar_archivos()
    elif opcion == '5':
        probar_subida_basica()
        probar_subida_excel()
        probar_crear_carpeta()
        probar_listar_archivos()
    elif opcion == '0':
        print("\n👋 ¡Hasta luego!\n")
        return False
    else:
        print("\n❌ Opción inválida")
    
    return True


if __name__ == "__main__":
    # Verificar archivos necesarios antes de empezar
    if not verificar_archivos_necesarios():
        print("\n⚠️  No se puede continuar sin los archivos necesarios")
        exit(1)
    
    print(f"\n{CONFIG['icono']} DOGEHOOT - PRUEBA DE GOOGLE DRIVE UPLOADER")
    print(f"🌍 Entorno: {CONFIG['entorno']}")
    
    if es_pythonanywhere():
        print("\n✅ Ejecutando en PythonAnywhere")
        print("   → Los tokens ya deben estar configurados")
    else:
        print("\n💻 Ejecutando en PC Local")
        print("   → La primera vez se abrirá un navegador para autenticarte")
        print("   → Se creará token_drive.json automáticamente")
        print("   → ¡Recuerda subir token_drive.json a PythonAnywhere después!")
    
    print()
    
    continuar = True
    while continuar:
        continuar = menu_principal()
