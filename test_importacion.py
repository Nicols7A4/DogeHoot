"""
Script de prueba para verificar que la importación funciona correctamente
"""
import openpyxl

# Crear un Excel de prueba
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Preguntas'

# Encabezados
headers = ['pregunta', 'opcion_1', 'opcion_2', 'opcion_3', 'opcion_4', 'opcion_c_orrecta', 'tiempo_segundos', 'url_imagen']
for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num, value=header)

# Datos de prueba - 3 preguntas
test_data = [
    ['¿Cuál es la capital de Francia?', 'París', 'Londres', 'Berlín', 'Madrid', 1, 15, ''],
    ['¿Cuánto es 5 x 5?', '20', '25', '30', '15', 2, 20, ''],
    ['¿Qué lenguaje se usa para web?', 'Python', 'JavaScript', 'C++', 'Java', 2, 30, ''],
]

for row_num, data in enumerate(test_data, start=2):
    for col_num, value in enumerate(data, 1):
        ws.cell(row=row_num, column=col_num, value=value)

# Guardar
wb.save('test_preguntas.xlsx')
print('✅ Archivo de prueba creado: test_preguntas.xlsx')
print('📝 Contiene 3 preguntas de prueba')
print('⚠️  Úsalo para probar la importación en el sistema')
